"""
LIFFアプリからのAPIリクエストを処理するBlueprint
"""
from flask import Blueprint, request, abort, jsonify, g, current_app, Response, make_response
from google.cloud import firestore
from google.cloud.firestore_v1.base_query import FieldFilter
import os
import sys
import csv
import io
import base64
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import requests
import json
import re
import uuid
from decimal import Decimal, ROUND_HALF_UP

from datetime import datetime, timezone, timedelta
from app_core.utils import get_user_info_by_line_id, get_calendar_statuses, get_all_category_b_labels, update_category_b_statuses, create_new_category_b, reactivate_category_b, check_unmapped_jobcan_employees, create_employee_mapping, calculate_monthly_period, update_category_b_offices, update_category_b_details, update_user_selection_history, get_user_selection_history, get_accommodation_notes_for_employees, get_on_site_status_for_employees, save_jobcan_holiday_types_to_firestore, enrich_holiday_types_payload_with_minutes, resolve_paid_leave_minutes_engineering, resolve_paid_leave_for_sync, resolve_paid_leave_for_sync_by_work_kind, is_net_main_group, default_net_paid_leave_time_slot
from app_core.monthly_closing_snapshot_selection import (
    classify_for_previous_month_closing_snapshot,
    compute_previous_monthly_period,
)
from app_core.utils import send_push_message, activate_download_link
# from app_core.utils import send_quick_report_email # 【実装保留】のためコメントアウト
from app_core.config import (
    CAREER_COEFFICIENT,
    COLLECTION_DAILY_REPORTS,
    COLLECTION_JOBCAN_RAW_RESPONSES,
    DAILY_REPORTS_SNAPSHOT_COLLECTION,
    MONTHLY_lABOR_COSTS_NET,
    MONTHLY_CLOSINGS_COLLECTION,
    default_snapshot_collection_for_closing_run,
    is_monthly_closing_test_mode,
    monthly_closings_collection_for_closing_run,
)
from functools import wraps
from flask_cors import CORS

# --- Blueprintの作成 ---
api_bp = Blueprint('api', __name__)
CORS(api_bp, supports_credentials=True)

# --- クライアントと定数の初期化 ---
db = firestore.Client()
import jwt
LINE_LOGIN_CHANNEL_ID = os.environ.get("LINE_LOGIN_CHANNEL_ID")
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN")
TEST_LINE_USER_ID = os.environ.get("TEST_LINE_USER_ID")
LINE_JWKS_URL = "https://api.line.me/oauth2/v2.1/certs"

# --- 認証ヘルパー関数 & デコレータ ---

def verify_line_id_token(id_token: str) -> dict:
    """JWKSを使用してLINEのIDトークンを検証し、デコードされたペイロードを返す。"""
    try:
        jwks_client = jwt.PyJWKClient(LINE_JWKS_URL)
        signing_key = jwks_client.get_signing_key_from_jwt(id_token)
        decoded_payload = jwt.decode(
            id_token,
            signing_key.key,
            algorithms=["ES256"],
            audience=LINE_LOGIN_CHANNEL_ID,
        )
        return decoded_payload
    except jwt.ExpiredSignatureError:
        abort(401, "トークンの有効期限が切れています。再度ログインしてください。")
    except jwt.InvalidAudienceError:
        abort(401, "トークンのオーディエンスが無効です。")
    except jwt.InvalidTokenError as e:
        abort(401, f"無効なトークンです: {e}")
    except jwt.PyJWTError as e:
        # abort(401)を投げることで、Flaskのエラーハンドラが適切なレスポンスを返す
        abort(401, f"IDトークンの検証に失敗しました: {e}")
    except Exception as e:
        # 予期せぬエラーは500として処理
        abort(500, f"An unexpected error occurred during token verification: {e}")

PC_SESSION_COOKIE_NAME = "dr_pc_session"
PC_SESSION_TTL_DAYS = 30
# NOTE: Cookie で使うため Secure は必須（SameSite=None の要件）
PC_SESSION_COOKIE_SAMESITE = "None"

def _get_pc_session_secret() -> str:
    secret = os.environ.get("PC_SESSION_JWT_SECRET")
    if not secret:
        current_app.logger.critical("PC_SESSION_JWT_SECRET environment variable is not set!")
        abort(500, "Server configuration error: PC session secret is not set.")
    return secret


def _build_period_key(start_date: datetime, end_date: datetime) -> str:
    return f"{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}"


def _load_monthly_closing_doc(
    period_key: str,
    division: str,
    *,
    closings_collection: str | None = None,
) -> dict | None:
    coll = closings_collection or MONTHLY_CLOSINGS_COLLECTION
    doc_id = f"{period_key}_{division}"
    doc = db.collection(coll).document(doc_id).get()
    if not doc.exists:
        return None
    data = doc.to_dict() or {}
    data["_doc_id"] = doc_id
    return data


def _is_monthly_closing_completed(
    period_key: str,
    division: str,
    *,
    closings_collection: str | None = None,
    default_snapshot_collection: str | None = None,
) -> tuple[bool, str]:
    """
    closings_collection / default_snapshot_collection を省略したときは本番の
    monthly_closings / daily_reports_snapshot を前提とする（集計の参照切替用）。
    """
    coll = closings_collection or MONTHLY_CLOSINGS_COLLECTION
    default_snap = default_snapshot_collection or DAILY_REPORTS_SNAPSHOT_COLLECTION
    data = _load_monthly_closing_doc(period_key, division, closings_collection=coll)
    if not data:
        return False, default_snap

    raw_status = data.get("status", "")
    status = str(raw_status).strip().lower()
    snapshot_collection = str(data.get("snapshot_collection") or default_snap).strip() or default_snap
    return status == "completed", snapshot_collection


def _resolve_summary_source_collection(target_month: str, start_date: datetime, end_date: datetime, division: str) -> str:
    """
    前月度のみ管理ドキュメントを参照して日報ソースを切り替える。
    - current: 常に daily_reports
    - previous:
      - net: 常に daily_reports（ネット向け集計は締め後スナップショットを使わない）
      - enj: 該当 division が completed のときのみ snapshot
      - all: enj/net の両方 completed のときのみ snapshot
    """
    if target_month != "previous":
        return COLLECTION_DAILY_REPORTS

    if division == "net":
        return COLLECTION_DAILY_REPORTS

    period_key = _build_period_key(start_date, end_date)
    if division == "enj":
        completed, snapshot_collection = _is_monthly_closing_completed(period_key, division)
        return snapshot_collection if completed else COLLECTION_DAILY_REPORTS

    if division == "all":
        enj_completed, enj_snapshot_collection = _is_monthly_closing_completed(period_key, "enj")
        net_completed, net_snapshot_collection = _is_monthly_closing_completed(period_key, "net")
        if enj_completed and net_completed:
            return enj_snapshot_collection or net_snapshot_collection or DAILY_REPORTS_SNAPSHOT_COLLECTION
        return COLLECTION_DAILY_REPORTS

    return COLLECTION_DAILY_REPORTS


_SNAPSHOT_COLLECTION_NAME_RE = re.compile(r"^[A-Za-z0-9_-]{1,128}$")


def _validate_snapshot_collection_name(name: str) -> str:
    n = (name or "").strip()
    if not _SNAPSHOT_COLLECTION_NAME_RE.fullmatch(n):
        abort(500, "スナップショットコレクション名が不正です。")
    return n


def _monthly_closing_query_date_bounds(period_start: datetime, period_end: datetime) -> tuple[datetime, datetime]:
    """`date` フィールドの範囲クエリ用に日付境界へ正規化する。"""
    qs = period_start.replace(hour=0, minute=0, second=0, microsecond=0)
    qe = period_end.replace(hour=23, minute=59, second=59, microsecond=999999)
    return qs, qe


def _employee_display_name_from_mappings(company_employee_id: str) -> str:
    """
    employee_mappings（ドキュメントID = 社員ID）の name を返す。
    取得できなければ社員IDをそのまま返す。
    """
    cid = (company_employee_id or "").strip()
    if not cid:
        return ""
    try:
        snap = db.collection("employee_mappings").document(cid).get()
        if snap.exists:
            raw = (snap.to_dict() or {}).get("name")
            if raw is not None:
                name = str(raw).strip()
                if name:
                    return name
    except Exception as e:
        current_app.logger.warning(
            "monthly closing: employee_mappings lookup failed for %s: %s",
            cid,
            e,
        )
    return cid


def execute_monthly_closing(
    division: str,
    started_by: str | None,
    *,
    period_start: datetime,
    period_end: datetime,
    period_key: str,
    closings_collection: str,
    snapshot_collection: str,
) -> dict:
    """
    前月度の日報を `daily_reports` から列挙し、division ごとにスナップショットへコピーする。
    `daily_reports` は読み取りのみ（更新・削除しない）。
    完了済み・実行中は 409。テスト/本番は closings_collection / snapshot_collection で切替。
    """
    division = str(division or "").strip().lower()
    if division not in ("enj", "net"):
        abort(400, "division は 'enj' または 'net' を指定してください。")

    snapshot_collection = _validate_snapshot_collection_name(snapshot_collection)
    started_by_employee_id = (started_by or "").strip()
    started_by_name = _employee_display_name_from_mappings(started_by_employee_id)

    mgmt_doc_id = f"{period_key}_{division}"
    mgmt_ref = db.collection(closings_collection).document(mgmt_doc_id)

    existing = _load_monthly_closing_doc(period_key, division, closings_collection=closings_collection)
    if existing:
        st = str(existing.get("status", "")).strip().lower()
        if st == "completed":
            abort(409, "この前月度・部署の締め処理はすでに完了しています。")
        if st == "running":
            abort(409, "この前月度・部署の締め処理が実行中です。完了後に再試行するか、システム管理者へ連絡してください。")

    other_div = "net" if division == "enj" else "enj"
    other = _load_monthly_closing_doc(period_key, other_div, closings_collection=closings_collection)
    if other and str(other.get("status", "")).strip().lower() == "running":
        abort(409, "他部署の締め処理が実行中です。完了後に再試行してください。")

    run_id = str(uuid.uuid4())
    mgmt_ref.set(
        {
            "status": "running",
            "division": division,
            "period_key": period_key,
            "period_start": period_start,
            "period_end": period_end,
            "snapshot_collection": snapshot_collection,
            "started_by": started_by_name,
            "started_by_employee_id": started_by_employee_id,
            "started_at": firestore.SERVER_TIMESTAMP,
            "run_id": run_id,
            "retry_allowed": False,
            "copied_count": 0,
        },
        merge=False,
    )

    q_start, q_end = _monthly_closing_query_date_bounds(period_start, period_end)
    try:
        query = (
            db.collection(COLLECTION_DAILY_REPORTS)
            .where(filter=FieldFilter("date", ">=", q_start))
            .where(filter=FieldFilter("date", "<=", q_end))
        )
        candidates = list(query.stream())
    except Exception as e:
        current_app.logger.exception("monthly closing: query daily_reports failed")
        mgmt_ref.update(
            {
                "status": "failed",
                "finished_at": firestore.SERVER_TIMESTAMP,
                "error_message": str(e)[:2000],
            }
        )
        abort(500, "日報の取得に失敗しました。")

    to_copy: list[tuple[str, dict]] = []
    for doc in candidates:
        data = doc.to_dict() or {}
        ok, div = classify_for_previous_month_closing_snapshot(
            data.get("date"),
            data.get("group_id"),
            period_start,
            period_end,
        )
        if ok and div == division:
            to_copy.append((doc.id, data))

    snap = db.collection(snapshot_collection)
    batch = db.batch()
    in_batch = 0
    committed_count = 0
    try:
        for doc_id, data in to_copy:
            batch.set(snap.document(doc_id), data, merge=False)
            in_batch += 1
            if in_batch >= 400:
                batch.commit()
                committed_count += in_batch
                batch = db.batch()
                in_batch = 0
        if in_batch:
            batch.commit()
            committed_count += in_batch
    except Exception as e:
        current_app.logger.exception("monthly closing: snapshot batch write failed")
        mgmt_ref.update(
            {
                "status": "failed",
                "finished_at": firestore.SERVER_TIMESTAMP,
                "error_message": str(e)[:2000],
                "copied_count_before_failure": committed_count,
            }
        )
        abort(500, "スナップショットへのコピーに失敗しました。")

    copied_count = committed_count
    mgmt_ref.update(
        {
            "status": "completed",
            "finished_at": firestore.SERVER_TIMESTAMP,
            "copied_count": copied_count,
        }
    )

    return {
        "status": "completed",
        "message": f"締め処理が完了しました（{copied_count} 件をスナップショットへコピー）。",
        "division": division,
        "started_by": started_by_name,
        "started_by_employee_id": started_by_employee_id,
        "updated": True,
        "copied_count": copied_count,
        "run_id": run_id,
    }


def _minimize_user_info_for_session(user_info: dict) -> dict:
    """
    Cookie サイズを抑えるため、認証・識別に必要な最小限のみ保持する。
    history などの肥大要素はセッションに入れない（必要時は別APIで取得）。
    """
    if not user_info or not isinstance(user_info, dict):
        return {}
    allowed_keys = [
        "company_employee_id",
        "jobcan_employee_id",
        "main_group_id",
        "name",
        "is_manager",
        "is_executive",
        "is_system_admin",
        "main_group_name",
        "mail",
        "work_kind_id",
    ]
    return {k: user_info.get(k) for k in allowed_keys}

def _encode_pc_session_jwt(line_user_id: str, user_info: dict) -> str:
    now = datetime.now(timezone.utc)
    exp = now + timedelta(days=PC_SESSION_TTL_DAYS)
    payload = {
        "typ": "dr_pc_session",
        "sub": str(line_user_id),
        "iat": int(now.timestamp()),
        "exp": int(exp.timestamp()),
        "user": _minimize_user_info_for_session(user_info),
    }
    return jwt.encode(payload, _get_pc_session_secret(), algorithm="HS256")

def _decode_pc_session_jwt(token: str) -> dict:
    try:
        payload = jwt.decode(
            token,
            _get_pc_session_secret(),
            algorithms=["HS256"],
            options={"require": ["exp", "iat", "sub"]},
        )
        if not isinstance(payload, dict) or payload.get("typ") != "dr_pc_session":
            abort(401, "セッションが無効です。")
        return payload
    except jwt.ExpiredSignatureError:
        abort(401, "セッションの有効期限が切れています。")
    except jwt.InvalidTokenError as e:
        abort(401, f"セッションが無効です: {e}")

def _try_authenticate_from_pc_session_cookie() -> bool:
    """
    PC用のセッションCookieがあれば検証して g に格納する。
    成功: True / Cookie無し or 失敗: False
    """
    token = request.cookies.get(PC_SESSION_COOKIE_NAME)
    if not token:
        return False
    payload = _decode_pc_session_jwt(token)
    line_user_id = payload.get("sub")
    user_info = payload.get("user")
    if not line_user_id or not isinstance(user_info, dict):
        abort(401, "セッションが無効です。")
    g.line_user_id = str(line_user_id)
    g.user_info = user_info
    g.auth_via_pc_session = True
    return True

def _authenticate_from_line_id_token_header() -> None:
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        abort(401, "Authorization header is missing or invalid.")
    try:
        id_token = auth_header.split(" ")[1]
        payload = verify_line_id_token(id_token)
        g.line_user_id = payload["sub"]
        g.auth_via_pc_session = False
    except (ValueError, IndexError, KeyError):
        abort(401, "ID token is invalid.")

def line_token_required(f):
    """LINE IDトークン（Authorization: Bearer ...）のみを検証する（PCセッションCookieは使わない）"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        _authenticate_from_line_id_token_header()
        return f(*args, **kwargs)
    return decorated_function

def token_required(f):
    """
    認証デコレータ
    - PC: セッションCookie（30日）を優先
    - スマホLIFF等: 従来通り LINE IDトークン（Bearer）
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if _try_authenticate_from_pc_session_cookie():
            return f(*args, **kwargs)
        _authenticate_from_line_id_token_header()
        return f(*args, **kwargs)
    return decorated_function

def manager_required(f):
    """
    ユーザーが管理者権限を持っていることを要求するデコレータ。
    @token_required の後に使用する必要がある。
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not hasattr(g, 'line_user_id'):
            abort(401, "Authentication token is required before checking manager status.")
        
        try:
            # PCセッション経由ならDBに触れずに権限判定
            user_info = g.user_info if hasattr(g, "user_info") and isinstance(g.user_info, dict) else get_user_info_by_line_id(g.line_user_id)
            if not user_info.get('is_manager'):
                abort(403, "Administrator privileges are required for this operation.")
            
            # ユーザー情報をgオブジェクトに格納して後続処理で利用可能にする
            g.user_info = user_info

        except Exception as e:
            # get_user_info_by_line_id が abort を送出する可能性がある
            if hasattr(e, 'code'):
                abort(e.code, e.description)
            abort(500, "An error occurred while verifying user permissions.")
            
        return f(*args, **kwargs)
    return decorated_function

def login_required(f):
    """
    ユーザーがシステムに登録済みであることを要求するデコレータ。
    @token_required の後に使用する。
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not hasattr(g, 'line_user_id'):
            abort(401, "Authentication token is required.")
        
        try:
            # PCセッション経由ならDBに触れずにユーザー識別
            if not (hasattr(g, "user_info") and isinstance(g.user_info, dict) and g.user_info.get("company_employee_id")):
                # ユーザー情報を取得 (未登録なら404エラーでabortされる)
                g.user_info = get_user_info_by_line_id(g.line_user_id)
        except Exception as e:
            if hasattr(e, 'code'):
                abort(e.code, e.description)
            abort(500, "An error occurred while verifying user registration.")
            
        return f(*args, **kwargs)
    return decorated_function


@api_bp.route("/manager/monthly-closing", methods=["POST"])
@token_required
@login_required
@manager_required
def post_manager_monthly_closing():
    """
    前月度の締め処理 API（第一段階）。
    管理ドキュメント上で当該 `period_key` + `division` が **completed** のときは 409 を返す。
    未完了のときは前月度の対象日報をスナップショットへコピーし、管理ドキュメントを completed にする。
    """
    data = request.get_json() or {}
    division = str(data.get("division", "")).strip().lower()
    if division not in ("enj", "net"):
        abort(400, "division は 'enj' または 'net' を指定してください。")

    period_start, period_end = compute_previous_monthly_period()
    period_key = _build_period_key(period_start, period_end)

    closings_coll = monthly_closings_collection_for_closing_run()
    default_snap = default_snapshot_collection_for_closing_run()
    completed, snapshot_collection = _is_monthly_closing_completed(
        period_key,
        division,
        closings_collection=closings_coll,
        default_snapshot_collection=default_snap,
    )
    if completed:
        abort(409, "この前月度・部署の締め処理はすでに完了しています。")

    if is_monthly_closing_test_mode():
        current_app.logger.warning(
            "MONTHLY_CLOSING_TEST_MODE is enabled: closing run uses "
            f"closings_collection={closings_coll!r} snapshot_collection={snapshot_collection!r}"
        )

    started_by = ""
    if hasattr(g, "user_info") and isinstance(g.user_info, dict):
        started_by = str(g.user_info.get("company_employee_id") or "").strip()

    payload = execute_monthly_closing(
        division,
        started_by,
        period_start=period_start,
        period_end=period_end,
        period_key=period_key,
        closings_collection=closings_coll,
        snapshot_collection=snapshot_collection,
    )
    payload["period_key"] = period_key
    payload["period_start"] = period_start.isoformat()
    payload["period_end"] = period_end.isoformat()
    payload["test_mode"] = is_monthly_closing_test_mode()
    payload["monthly_closings_collection"] = closings_coll
    payload["snapshot_collection"] = snapshot_collection
    return jsonify(payload), 200


def _jsonable_firestore_scalar(value):
    """Firestore ドキュメント値を JSON 化しやすい形へ（タイムスタンプ等）。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "timestamp") and callable(getattr(value, "timestamp")):
        try:
            return datetime.fromtimestamp(value.timestamp(), tz=timezone.utc).isoformat()
        except Exception:
            pass
    return str(value)


def _monthly_closing_division_status_for_dashboard(period_key: str, division: str) -> dict:
    """ダッシュボード用。常に本番 `MONTHLY_CLOSINGS_COLLECTION` のみ参照する。"""
    data = _load_monthly_closing_doc(period_key, division, closings_collection=MONTHLY_CLOSINGS_COLLECTION)
    if not data:
        return {"exists": False, "status": None}

    st = str(data.get("status", "")).strip().lower() or None
    out: dict = {
        "exists": True,
        "status": st,
        "finished_at": _jsonable_firestore_scalar(data.get("finished_at")),
        "started_at": _jsonable_firestore_scalar(data.get("started_at")),
        "started_by": data.get("started_by"),
        "started_by_employee_id": data.get("started_by_employee_id"),
        "copied_count": data.get("copied_count"),
        "run_id": data.get("run_id"),
    }
    if st == "failed":
        out["error_message"] = data.get("error_message")
    return out


@api_bp.route("/manager/monthly-closing/status", methods=["GET"])
@token_required
@login_required
@manager_required
def get_manager_monthly_closing_status():
    """
    ダッシュボード用の締め状態（前月度）。
    **常に本番** `monthly_closings` のみを参照する（MONTHLY_CLOSING_TEST_MODE の影響を受けない）。
    """
    period_start, period_end = compute_previous_monthly_period()
    period_key = _build_period_key(period_start, period_end)
    closing_day = int(os.environ.get("CLOSING_DAY", "20"))
    period_end_date = period_end.date().isoformat() if hasattr(period_end, "date") else str(period_end)[:10]

    payload = {
        "period_key": period_key,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "period_end_date": period_end_date,
        "closing_day": closing_day,
        "closings_collection": MONTHLY_CLOSINGS_COLLECTION,
        "enj": _monthly_closing_division_status_for_dashboard(period_key, "enj"),
        "net": _monthly_closing_division_status_for_dashboard(period_key, "net"),
    }
    return jsonify(payload), 200


def get_authenticated_user_info() -> dict:
    """
    認証済みユーザー情報を取得する。
    - PCセッションCookie経由: g.user_info をそのまま使用（DBアクセスなし）
    - それ以外: Firestore から取得し g.user_info にキャッシュ
    """
    if hasattr(g, "user_info") and isinstance(g.user_info, dict) and g.user_info.get("company_employee_id"):
        return g.user_info
    if not hasattr(g, "line_user_id"):
        abort(401, "Authentication token is required.")
    g.user_info = get_user_info_by_line_id(g.line_user_id)
    return g.user_info

@api_bp.route("/pc/session", methods=["POST"])
@line_token_required
def create_pc_session():
    """
    PC初回ログイン用:
    - LINE IDトークンを検証
    - Firestore からユーザー情報を1回だけ取得
    - 30日有効のセッションCookie（署名付きJWT）を発行
    """
    line_user_id = g.line_user_id
    user_info = get_user_info_by_line_id(line_user_id)
    token = _encode_pc_session_jwt(line_user_id, user_info)

    resp = make_response(jsonify({"status": "ok"}), 200)
    resp.set_cookie(
        PC_SESSION_COOKIE_NAME,
        token,
        max_age=int(PC_SESSION_TTL_DAYS * 24 * 60 * 60),
        httponly=True,
        secure=True,
        samesite=PC_SESSION_COOKIE_SAMESITE,
        path="/",
    )
    return resp

@api_bp.route("/pc/session", methods=["DELETE"])
def delete_pc_session():
    """PCセッションを破棄（ログアウト用）"""
    resp = make_response(jsonify({"status": "ok"}), 200)
    resp.set_cookie(
        PC_SESSION_COOKIE_NAME,
        "",
        max_age=0,
        httponly=True,
        secure=True,
        samesite=PC_SESSION_COOKIE_SAMESITE,
        path="/",
    )
    return resp

def scheduler_token_required(f):
    """Cloud Schedulerからのリクエストを認証するデコレータ"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Cloud Schedulerからのリクエストには、このヘッダーが含まれる
        # ローカルでのテスト用に、ヘッダーがない場合も許容する（環境変数で制御）
        if os.environ.get("FLASK_ENV") != "development" and not request.headers.get("X-CloudScheduler"):
             abort(403, "Forbidden: Not a Cloud Scheduler request.")

        auth_header = request.headers.get("Authorization")
        scheduler_token = os.environ.get('SCHEDULER_TOKEN')
        if not scheduler_token:
            current_app.logger.critical("SCHEDULER_TOKEN environment variable is not set!")
            abort(500, "Server configuration error: Scheduler token is not set.")

        expected_token = f"Bearer {scheduler_token}"
        if not auth_header or auth_header != expected_token:
            current_app.logger.error(f"Scheduler authentication failed. Received: '{auth_header}', Expected: '{expected_token}'")
            abort(401, "Unauthorized: Invalid scheduler token.")
        return f(*args, **kwargs)
    return decorated_function
# --- APIエンドポイント定義 ---

@api_bp.route("/verify-employee", methods=["POST"])
@token_required
def verify_employee():
    """LIFFアプリから送られた社員IDを検証し、名前を返すエンドポイント"""
    data = request.get_json()
    if not data or "employeeId" not in data:
        abort(400, "Request body must contain 'employeeId'.")
    employee_id = data["employeeId"]

    mapping_ref = db.collection("employee_mappings").document(employee_id)
    mapping_doc = mapping_ref.get()

    if not mapping_doc.exists:
        abort(404, f"社員ID「{employee_id}」は見つかりませんでした。")

    mapping_data = mapping_doc.to_dict()
    name = mapping_data.get("name", "名前未登録")
    return jsonify({"status": "found", "name": name, "employeeId": employee_id}), 200

@api_bp.route("/register", methods=["POST"])
@token_required
def register_from_liff():
    """LIFFアプリから社員ID登録を受け付けるエンドポイント"""
    user_id = g.line_user_id
    data = request.get_json()
    if not data or "employeeId" not in data:
        abort(400, "Request body must contain 'employeeId'.")
    
    employee_id = data["employeeId"]
    if not isinstance(employee_id, str) or not employee_id.isdigit():
        abort(400, "'employeeId' must be a string containing only digits.")

    # 1. 登録しようとしている社員IDがemployee_mappingsに存在するか確認
    mapping_ref = db.collection("employee_mappings").document(employee_id)
    mapping_doc = mapping_ref.get()
    if not mapping_doc.exists:
        abort(404, f"社員ID「{employee_id}」はシステムに登録されていません。")
    
    mapping_data = mapping_doc.to_dict()
    jobcan_employee_id = mapping_data.get("jobcan_employee_id")
    is_manager = mapping_data.get("is_manager", False)
    mail = mapping_data.get("mail") # mailフィールドも取得

    # Jobcanからメイングループを取得して保存する
    main_group_id = None
    if jobcan_employee_id:
        try:
            from services.jobcan_service import JobcanService
            app_env = os.environ.get("APP_ENV", "development")
            is_sandbox = app_env != "production"
            jobcan_service = JobcanService(sandbox=is_sandbox)
            employee_details = jobcan_service.get_employee_details(employee_code=jobcan_employee_id)
            if employee_details:
                main_group_id = employee_details.get("main_group")
        except Exception as e:
            current_app.logger.warning(f"Failed to fetch main_group during registration: {e}")

    # 2. このLINEアカウントが既に他の社員IDに紐付いていないか確認
    users_ref = db.collection("users")
    query = users_ref.where(filter=FieldFilter("line_user_id", "==", user_id)).limit(1)
    docs = list(query.stream())
    if docs:
        existing_doc = docs[0]
        linked_company_id = existing_doc.id
        if linked_company_id != employee_id:
            abort(409, f"このLINEアカウントは既に別の社員ID（{linked_company_id}）に紐付いています。")

    # 3. usersコレクションにデータを書き込む
    user_ref = users_ref.document(employee_id)
    user_doc = user_ref.get()

    user_data_to_set = {
        "jobcan_employee_id": jobcan_employee_id,
        "line_user_id": user_id,
        "is_manager": is_manager,
        "mail": mail, # mailフィールドを保存
        "main_group": main_group_id, # メイングループIDを保存
        "updated_at": firestore.SERVER_TIMESTAMP
    }

    if not user_doc.exists:
        user_data_to_set["created_at"] = firestore.SERVER_TIMESTAMP
        user_ref.set(user_data_to_set)
        action = "created"
    else:
        user_ref.update(user_data_to_set)
        action = "updated"

    return jsonify({"status": action, "employeeId": employee_id}), 200


@api_bp.route("/reports", methods=["POST"])
@token_required
def post_report():
    """LIFFアプリや管理画面から業務報告を受け取り、Firestoreに保存するエンドポイント"""
    user_id = g.line_user_id

    report_data = request.get_json()
    if not report_data:
        abort(400, "Request body is missing or not a valid JSON.")

    date = report_data.get("date")
    work_time = report_data.get("taskTotalMinutes")
    jobcan_work_minutes = report_data.get("jobcanWorkMinutes")
    tasks = report_data.get("tasks")

    # 代理入力用のパラメータを取得
    is_proxy = report_data.get("is_proxy", False)
    target_employee_id_from_req = report_data.get("target_employee_id")

    if not date or work_time is None or jobcan_work_minutes is None:
        abort(400, "Invalid request body. 'date', 'taskTotalMinutes', and 'jobcanWorkMinutes' are required.")

    # 受信値はバリデーション用途のみ。保存値は tasks 合計で再計算して確定する。
    try:
        requested_total = int(work_time)
    except (TypeError, ValueError):
        requested_total = 0

    if tasks is None or (requested_total > 0 and not tasks):
        abort(400, "Invalid request body. 'tasks' is required when 'taskTotalMinutes' is greater than 0.")

    report_content_lines = [
        f"【{task.get('time', 0)}分】{task.get('categoryA_label', '')} - {task.get('categoryB_label', '')}"
        for task in (tasks or [])
    ]
    report_content = "\n".join(report_content_lines)
    # task_total_minutes はサーバー側で tasks から再計算して正とする
    work_time_minutes = 0
    for task in (tasks or []):
        if not isinstance(task, dict):
            continue
        try:
            work_time_minutes += int(task.get("time", 0) or 0)
        except (TypeError, ValueError):
            continue
    date_obj = datetime.strptime(date, '%Y-%m-%d')

    # --- 報告者と入力者の情報を決定 ---
    inputter_info = get_user_info_by_line_id(user_id)
    inputter_id = inputter_info["company_employee_id"]
    inputter_name = inputter_info.get("name")

    target_employee_id = None
    target_employee_name = None
    target_group_id = None
    target_group_name = None

    if is_proxy and target_employee_id_from_req:
        # 代理入力の場合
        target_employee_id = target_employee_id_from_req
        try:
            # 報告対象者の情報を取得 (usersとemployee_mappingsから)
            target_user_ref = db.collection("users").document(target_employee_id)
            target_user_doc = target_user_ref.get()
            if not target_user_doc.exists:
                abort(404, f"報告対象のユーザー(ID: {target_employee_id})が見つかりません。")
            
            target_user_data = target_user_doc.to_dict()
            target_group_id = target_user_data.get("main_group")
            
            target_mapping_ref = db.collection("employee_mappings").document(target_employee_id)
            target_mapping_doc = target_mapping_ref.get()
            target_employee_name = target_mapping_doc.to_dict().get("name") if target_mapping_doc.exists else "名前不明"

            if target_group_id:
                group_mapping_ref = db.collection("group_mappings").document(str(target_group_id))
                group_mapping_doc = group_mapping_ref.get()
                target_group_name = group_mapping_doc.to_dict().get("name") if group_mapping_doc.exists else "グループ名不明"
            else:
                target_group_name = "（未設定）"

        except Exception as e:
            current_app.logger.error(f"Failed to get target user info for proxy report: {e}")
            abort(500, "報告対象のユーザー情報の取得に失敗しました。")
    else:
        # 本人入力の場合
        target_employee_id = inputter_id
        target_employee_name = inputter_name
        target_group_id = inputter_info.get("main_group_id")
        target_group_name = inputter_info.get("main_group_name")
        is_proxy = False

    doc_id = f"{target_employee_id}_{date}"

    try:
        doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
        doc_ref.set({
            # --- 集計用に追加するフィールド ---
            "employee_name": target_employee_name, # 報告対象者
            "inputter_id": inputter_id,            # 入力者の社員ID
            "inputter_name": inputter_name,        # 入力者名
            "is_proxy_report": is_proxy,           # 代理報告フラグ
            "group_id": target_group_id,
            "group_name": target_group_name,
            "report_year": date_obj.year,
            "report_month": date_obj.month,
            "jobcan_work_minutes": int(jobcan_work_minutes),
            # --- 既存のフィールド ---
            "date": date_obj,
            "company_employee_id": target_employee_id,
            "task_total_minutes": work_time_minutes,
            "tasks": tasks,
            "report_content": report_content,
            "report_updated_at": firestore.SERVER_TIMESTAMP
        })
    except Exception as e:
        print(f"Error updating Firestore: {e}")
        abort(500, "Failed to save report.")

    # --- 履歴の更新 ---
    # 報告対象者（target_employee_id）の履歴を更新する
    update_user_selection_history(target_employee_id, tasks)

    return jsonify({"status": "success"}), 200

@api_bp.route("/reports_net", methods=["POST"])
@token_required
def post_net_report():
    """
    【ネット事業部専用】タイムテーブル形式（開始・終了時刻付き）の業務報告を受け取り、Firestoreに保存する。
    既存の /api/reports とは独立して動作する。
    """
    user_id = g.line_user_id

    report_data = request.get_json()
    if not report_data:
        abort(400, "Request body is missing or not a valid JSON.")

    # --- 新しい形式のデータをパース ---
    date = report_data.get("date")
    work_time = report_data.get("taskTotalMinutes")
    jobcan_work_minutes = report_data.get("jobcanWorkMinutes")
    tasks_from_req = report_data.get("tasks") # startTime, endTime, comment を含むタスク

    # 代理入力用のパラメータを取得
    is_proxy = report_data.get("is_proxy", False)
    target_employee_id_from_req = report_data.get("target_employee_id")

    if not date or work_time is None or jobcan_work_minutes is None:
        abort(400, "Invalid request body. 'date', 'taskTotalMinutes', and 'jobcanWorkMinutes' are required.")

    # 受信値はバリデーション用途のみ。保存値は tasks 合計で再計算して確定する。
    try:
        requested_total = int(work_time)
    except (TypeError, ValueError):
        requested_total = 0

    if tasks_from_req is None or (requested_total > 0 and not tasks_from_req):
        abort(400, "Invalid request body. 'tasks' is required when 'taskTotalMinutes' is greater than 0.")

    # --- 新しい形式に合わせて report_content と tasks を生成 ---
    report_content_lines = []
    tasks_to_save = []
    for task in (tasks_from_req or []):
        # 必須フィールドのチェック
        if not all(k in task for k in ['categoryA_label', 'categoryB_label', 'time', 'startTime', 'endTime']):
            continue # 不完全なタスクはスキップ

        # report_content の生成
        report_content_lines.append(
            f"【{task.get('startTime')}~{task.get('endTime')}|{task.get('time', 0)}分】{task.get('categoryA_label', '')} - {task.get('categoryB_label', '')}"
        )
        
        # 保存用タスクオブジェクトの生成
        tasks_to_save.append({
            "categoryA_id": task.get("categoryA_id"),
            "categoryA_label": task.get("categoryA_label"),
            "categoryB_id": task.get("categoryB_id"),
            "categoryB_label": task.get("categoryB_label"),
            "time": task.get("time"),
            "startTime": task.get("startTime"), # 新しいフィールド
            "endTime": task.get("endTime"),     # 新しいフィールド
            "comment": task.get("comment", "")  # 新しいフィールド
        })

    report_content = "\n".join(report_content_lines)
    # task_total_minutes はサーバー側で tasks から再計算して正とする
    work_time_minutes = 0
    for task in tasks_to_save:
        if not isinstance(task, dict):
            continue
        try:
            work_time_minutes += int(task.get("time", 0) or 0)
        except (TypeError, ValueError):
            continue
    
    # 文字列の日付をdatetimeオブジェクトに変換。FirestoreはこれをTimestampとして保存する。
    date_obj = datetime.strptime(date, '%Y-%m-%d')

    # --- 報告者と入力者の情報を決定 (post_reportから流用) ---
    inputter_info = get_authenticated_user_info()
    inputter_id = inputter_info["company_employee_id"]
    inputter_name = inputter_info.get("name")

    target_employee_id = None
    target_employee_name = None
    target_group_id = None
    target_group_name = None

    if is_proxy and target_employee_id_from_req:
        # 代理入力の場合
        target_employee_id = target_employee_id_from_req
        try:
            target_user_ref = db.collection("users").document(target_employee_id)
            target_user_doc = target_user_ref.get()
            if not target_user_doc.exists:
                abort(404, f"報告対象のユーザー(ID: {target_employee_id})が見つかりません。")
            
            target_user_data = target_user_doc.to_dict()
            target_group_id = target_user_data.get("main_group")
            
            target_mapping_ref = db.collection("employee_mappings").document(target_employee_id)
            target_mapping_doc = target_mapping_ref.get()
            target_employee_name = target_mapping_doc.to_dict().get("name") if target_mapping_doc.exists else "名前不明"

            if target_group_id:
                group_mapping_ref = db.collection("group_mappings").document(str(target_group_id))
                group_mapping_doc = group_mapping_ref.get()
                target_group_name = group_mapping_doc.to_dict().get("name") if group_mapping_doc.exists else "グループ名不明"
            else:
                target_group_name = "（未設定）"

        except Exception as e:
            current_app.logger.error(f"Failed to get target user info for proxy report: {e}")
            abort(500, "報告対象のユーザー情報の取得に失敗しました。")
    else:
        # 本人入力の場合
        target_employee_id = inputter_id
        target_employee_name = inputter_name
        target_group_id = inputter_info.get("main_group_id")
        target_group_name = inputter_info.get("main_group_name")
        is_proxy = False

    doc_id = f"{target_employee_id}_{date}"

    try:
        doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
        doc_ref.set({
            # --- 集計用に追加するフィールド ---
            "employee_name": target_employee_name,
            "inputter_id": inputter_id,
            "inputter_name": inputter_name,
            "is_proxy_report": is_proxy,
            "group_id": target_group_id,
            "group_name": target_group_name,
            "report_year": date_obj.year,
            "report_month": date_obj.month,
            "jobcan_work_minutes": int(jobcan_work_minutes),
            # --- 既存のフィールド ---
            "date": date_obj, # datetimeオブジェクトを渡す
            "company_employee_id": target_employee_id,
            "task_total_minutes": work_time_minutes,
            "tasks": tasks_to_save, # 新しい形式のタスク
            "report_content": report_content,
            "report_updated_at": firestore.SERVER_TIMESTAMP # サーバーサイドのタイムスタンプ
        })
    except Exception as e:
        print(f"Error updating Firestore: {e}")
        abort(500, "Failed to save report.")

    # --- 履歴の更新 (post_reportから流用) ---
    update_user_selection_history(target_employee_id, tasks_to_save)

    return jsonify({"status": "success"}), 201

@api_bp.route("/report-details", methods=["GET"])
@token_required
def get_report_details():
    """指定された日付の工数報告詳細を取得するエンドポイント。管理者はemployee_idを指定して他人の情報を取得可能。"""
    target_date = request.args.get("date")
    target_employee_id = request.args.get("employee_id") # 代理入力用

    if not target_date:
        abort(400, "Query parameter 'date' is required.")

    company_employee_id = None
    if target_employee_id:
        # 代理入力の場合、指定されたIDを使用
        company_employee_id = target_employee_id
    else:
        # 通常の本人入力の場合
        user_info = get_user_info_by_line_id(g.line_user_id)
        company_employee_id = user_info["company_employee_id"]

    # ドキュメントIDを構築してFirestoreからデータを取得
    doc_id = f"{company_employee_id}_{target_date}"
    doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
    doc = doc_ref.get()

    if doc.exists:
        return jsonify({**doc.to_dict(), "date": doc.to_dict()["date"].strftime('%Y-%m-%d')}), 200
    else:
        # 存在しなければ、空のタスクリストを返す
        return jsonify({"tasks": []}), 200

@api_bp.route("/send-announcement", methods=["POST"])
@token_required
@manager_required
def send_announcement():
    """
    全社連絡を送信する。
    is_test=True の場合は、環境変数 TEST_LINE_USER_ID で指定されたユーザーのみに送信する。
    """
    if not LINE_CHANNEL_ACCESS_TOKEN:
        return jsonify({"message": "Server Error: LINE Channel Access Token is not set."}), 500

    data = request.json
    message_text = data.get('message')
    sender_name = data.get('sender_name')
    is_test = data.get('is_test', False)

    if not message_text:
        return jsonify({"message": "Message is required"}), 400

    # メッセージ本文の作成
    full_message = f"{message_text}\n\n[{sender_name}より]"
    
    if is_test:
        # テスト送信: 特定のユーザーIDにプッシュメッセージ
        target_user_id = TEST_LINE_USER_ID
        if not target_user_id:
             return jsonify({"message": "Server Error: TEST_LINE_USER_ID is not set."}), 500

        # 環境変数から読み込んだ値に不要な空白や引用符が含まれている可能性を考慮
        target_user_id = target_user_id.strip().strip('"').strip("'")
        
        # アクセストークンもクリーニング
        clean_token = LINE_CHANNEL_ACCESS_TOKEN.strip()

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {clean_token}"
        }
        payload = {
            "to": target_user_id,
            "messages": [{"type": "text", "text": f"【テスト送信】\n{full_message}"}]
        }
        
        # デバッグ用ログ: 送信ペイロード（IDはマスク）
        masked_payload = payload.copy()
        masked_payload['to'] = target_user_id[:4] + '***' + target_user_id[-4:] if len(target_user_id) > 8 else '***'
        current_app.logger.info(f"Sending LINE Push: {masked_payload}")
        
        try:
            response = requests.post("https://api.line.me/v2/bot/message/push", headers=headers, json=payload)
            response.raise_for_status()
            return jsonify({"message": "Test message sent successfully"}), 200
        except requests.exceptions.RequestException as e:
            # LINE APIからの詳細なエラーレスポンスをログに出力する
            error_details = ""
            if e.response is not None:
                try:
                    # レスポンスボディがJSON形式の場合、整形してログに出力
                    error_details = e.response.json()
                except ValueError:
                    # JSONでなければテキストとして出力
                    error_details = e.response.text
            
            current_app.logger.error(f"LINE API Error: {e}, Details: {error_details}")
            # フロントエンドにも詳細を返す
            return jsonify({"message": f"Failed to send message to LINE API: {e}", "details": error_details}), 500
    else:
        # 本番送信 (ブロードキャスト) - まだ実装しない
        return jsonify({"message": "Broadcast is not implemented yet."}), 501

@api_bp.route("/user", methods=["GET"])
@token_required
def get_user_info():
    """
    現在のLINEユーザーに紐づく社員情報を返すエンドポイント。
    フロントエンドが報告者名を表示するために使用する。
    """
    # PCセッション経由ならDBに触れず、セッション内の情報のみで返す
    if hasattr(g, "auth_via_pc_session") and g.auth_via_pc_session and hasattr(g, "user_info") and isinstance(g.user_info, dict):
        user_info = g.user_info
        return jsonify({
            "employeeId": user_info.get("company_employee_id"),
            "name": user_info.get("name"),
            "is_manager": user_info.get("is_manager", False),
            "is_system_admin": user_info.get("is_system_admin", False),
            "main_group_name": user_info.get("main_group_name"),
            "main_group": user_info.get("main_group_id"),
            "is_executive": user_info.get("is_executive", False),
            "history": {"catA": [], "catB": []},
            "work_kind_id": user_info.get("work_kind_id"),
        }), 200

    # それ以外（スマホLIFF等）は従来通りDBから取得
    user_id = g.line_user_id
    user_info = get_user_info_by_line_id(user_id)

    # main_group が未設定の場合、Jobcanから取得して補完する
    if user_info.get("main_group_id") is None and user_info.get("jobcan_employee_id"):
        try:
            from services.jobcan_service import JobcanService
            app_env = os.environ.get("APP_ENV", "development")
            is_sandbox = app_env != "production"
            jobcan_service = JobcanService(sandbox=is_sandbox)
            
            employee_details = jobcan_service.get_employee_details(employee_code=user_info["jobcan_employee_id"])
            latest_group_id = employee_details.get("main_group") if employee_details else None
            
            if latest_group_id is not None:
                # Firestoreを更新
                db.collection("users").document(user_info["company_employee_id"]).update({
                    "main_group": latest_group_id,
                    "updated_at": firestore.SERVER_TIMESTAMP
                })
                # レスポンス用の辞書も更新
                user_info["main_group_id"] = latest_group_id
                
                # グループ名も取得して更新（表示用）
                group_doc = db.collection("group_mappings").document(str(latest_group_id)).get()
                if group_doc.exists:
                    user_info["main_group_name"] = group_doc.to_dict().get("name")
                    
        except Exception as e:
            current_app.logger.warning(f"Failed to fetch main_group from Jobcan in /user: {e}")

    # get_user_info_by_line_id はユーザーが見つからない場合にabort(404)を発生させるため、
    # ここでの追加のチェックは不要。

    # フロントエンドが期待するキー名 'employeeId' に変換して返す
    # ユーザーが見つからない場合は、get_user_info_by_line_idがabortするため、このコードには到達しない
    return jsonify({
        "employeeId": user_info.get("company_employee_id"),
        "name": user_info.get("name"),
        "is_manager": user_info.get("is_manager", False),
        "is_system_admin": user_info.get("is_system_admin", False),
        "main_group_name": user_info.get("main_group_name"), # 表示名を返す
        "main_group": user_info.get("main_group_id"), # グループIDを返す
        "is_executive": user_info.get("is_executive", False), # フロントエンドへ返す
        "history": user_info.get("history", {"catA": [], "catB": []}), # 履歴を返す
        # users に存在する場合のみ値あり（未同期・未設定は null）。参照先はまず本フィールドのみ。
        "work_kind_id": user_info.get("work_kind_id"),
    }), 200

@api_bp.route("/work-time", methods=["GET"])
@token_required
def get_work_time():
    """
    指定された日付の勤務時間をJobcanから取得して返すエンドポイント。
    管理者は employee_id クエリを指定することで他従業員の情報を取得可能。
    """
    user_id = g.line_user_id
    target_date = request.args.get("date")
    target_employee_id = request.args.get("employee_id") # 管理者用パラメータ
    source = request.args.get("source", "report") # デフォルトは 'report'
    # フロントエンドから待機時間を指定するためのパラメータ
    try:
        wait_param = float(request.args.get("wait", "0"))
    except (ValueError, TypeError):
        wait_param = 0.0

    if not target_date:
        abort(400, "Query parameter 'date' is required.")

    # 1. ユーザー情報と、対象となるjobcan_employee_idを取得
    user_info = get_authenticated_user_info()
    jobcan_employee_id = None
    target_company_id = None # Firestore更新用に会社IDを保持

    # 管理者が他従業員の情報を取得する場合
    if target_employee_id:
        # PC画面公開に伴い、他人の勤務時間取得制限を緩和
        # if not user_info.get('is_manager'):
        #     abort(403, "Administrator privileges are required to specify an employee_id.")
        
        target_company_id = target_employee_id # IDを設定

        # 指定された company_employee_id から jobcan_employee_id を取得
        mapping_ref = db.collection("employee_mappings").document(target_employee_id)
        mapping_doc = mapping_ref.get()
        if not mapping_doc.exists:
            current_app.logger.warning(f"Work time requested for non-existent employee mapping: {target_employee_id}")
            return jsonify({"date": target_date, "workTime": 0}), 200
        jobcan_employee_id = mapping_doc.to_dict().get("jobcan_employee_id")
    else:
        # 一般ユーザーが自身の情報を取得する場合
        jobcan_employee_id = user_info.get("jobcan_employee_id")
        target_company_id = user_info.get("company_employee_id") # IDを設定

    # Jobcan IDがない場合は0を返す
    if not jobcan_employee_id:
        return jsonify({"date": target_date, "workTime": 0}), 200

    # 当日かどうかを判定 (JST基準)
    today_str = (datetime.now(timezone(timedelta(hours=9)))).strftime('%Y-%m-%d')

    # 2. JobcanServiceを使って勤務サマリーを取得
    try:
        # config.pyへの依存をなくすため、services.jobcan_serviceは関数内でimportする
        from services.jobcan_service import JobcanService

        # 実行環境に応じてサンドボックスフラグを設定
        # Cloud Runの環境変数 APP_ENV が 'production' の場合は本番用、それ以外はサンドボックス用を使用
        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"

        jobcan_service = JobcanService(
            db=db,
            sandbox=is_sandbox, # JobcanServiceがこのフラグを元に認証情報とURLを決定する
            raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES
        )

        work_minutes = 0
        # 当日の場合は打刻情報から計算するメソッドを呼び出す
        if target_date == today_str:
            # 呼び出し元に応じて待機時間を設定
            # waitパラメータが指定されていればそれを優先する
            if wait_param > 0:
                wait_seconds = wait_param
            else:
                wait_seconds = 2.0 if source == 'report' else 0

            current_app.logger.info(
                f"Fetching real-time work time for today ({target_date}) from '{source}'. Wait: {wait_seconds}s"
            )
            work_minutes = jobcan_service.calculate_work_time_from_adits(employee_id=jobcan_employee_id, date=target_date, wait_seconds=wait_seconds)
        else:
            # 当日以外は従来通りサマリーAPIを呼び出す
            current_app.logger.info(f"Fetching summarized work time for {target_date}")
            
            # Jobcanの集計が完了していない場合('pending')のポーリング処理を追加
            max_retries = 5
            for i in range(max_retries):
                response_data = jobcan_service.get_daily_summaries(employee_id=jobcan_employee_id, dates=[target_date])
                
                if response_data and 'refresh' in response_data and response_data['refresh'].get('status') == 'pending':
                    if i < max_retries - 1:
                        current_app.logger.info(f"Jobcan summary refresh pending. Retrying in 2s... ({i+1}/{max_retries})")
                        time.sleep(2)
                        continue
                break

            if response_data and response_data.get("daily_summaries"):
                summary = response_data["daily_summaries"][0]
                work_minutes = summary.get("work", 0)
        
        # --- 取得した最新の勤務時間でFirestoreの日報データを更新する ---
        if target_company_id:
            doc_id = f"{target_company_id}_{target_date}"
            doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
            try:
                # 日報ドキュメントが存在する場合のみ、jobcan_work_minutesを更新する
                # update() はドキュメントが存在しない場合にNotFoundエラーを発生させる
                doc_ref.update({"jobcan_work_minutes": work_minutes})
            except Exception:
                # 日報が未提出（ドキュメントが存在しない）場合は何もしない
                pass
        
        return jsonify({"date": target_date, "workTime": work_minutes}), 200
    except Exception as e:
        print(f"Error getting work time from Jobcan: {e}")
        abort(500, "Failed to get work time from Jobcan.")

@api_bp.route("/calendar-statuses", methods=["GET"])
@token_required
def get_calendar_status_for_month():
    """
    指定された年月（または現在）の勤務状況ステータスを返すエンドポイント。
    フロントエンドがカレンダーの表示を更新するために使用する。
    """
    # クエリパラメータから 'start_date' と 'end_date' を取得
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    if not start_date_str or not end_date_str:
        abort(400, "Query parameters 'start_date' and 'end_date' are required.")

    # 【デバッグ】エンドポイントが呼び出されたことをログに出力
    current_app.logger.info(f"[/api/calendar-statuses] endpoint called with start: {start_date_str}, end: {end_date_str}")

    # ユーザー情報を取得
    user_info = get_authenticated_user_info()
    jobcan_employee_id = user_info.get("jobcan_employee_id")
    company_employee_id = user_info.get("company_employee_id")
    is_executive = user_info.get("is_executive", False)

    # 指定された期間の日付リストを生成
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        if start_date > end_date:
            abort(400, "'start_date' must be before or the same as 'end_date'.")
        dates_in_month = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end_date - start_date).days + 1)]
    except ValueError:
        abort(400, "Invalid date format. Please use YYYY-MM-DD.")

    statuses = get_calendar_statuses(jobcan_employee_id, company_employee_id, dates_in_month, is_executive=is_executive)
    return jsonify(statuses), 200

@api_bp.route("/manager/categories/b", methods=["GET"])
@token_required
@manager_required
def get_manager_category_b():
    """
    管理者向けに、category_bの一覧（詳細情報含む）を返すエンドポイント。
    """
    kind = request.args.get("kind")
    if not kind:
        abort(400, "Query parameter 'kind' is required.")

    # 'kind'の値が想定通りかバリデーション
    if kind not in ["engineering", "net"]:
        abort(400, "Invalid 'kind' parameter. Must be 'engineering' or 'net'.")

    try:
        query = db.collection("category_b").where(filter=FieldFilter("kind", "==", kind))
        docs = query.stream()
        
        categories = []
        for doc in docs:
            data = doc.to_dict()
            categories.append({
                "id": doc.id,
                "label": data.get("label", ""),
                "kind": data.get("kind", ""),
                "active": data.get("active", True),
                "order": data.get("order", 0),
                "client": data.get("client", ""),
                "project": data.get("project", ""),
                "offices": data.get("offices", []),
                "category_a_settings": data.get("category_a_settings", {}),  # color_map -> category_a_settings
                "category_a_sort": data.get("category_a_sort", {})  # PC入力画面専用 並び順
            })
        
        # order降順、label昇順でソート
        categories.sort(key=lambda x: (-x['order'], x['label']))
        
        return jsonify(categories), 200
    except Exception as e:
        current_app.logger.error(f"Failed to fetch categories: {e}")
        abort(500, f"Failed to fetch categories: {e}")


@api_bp.route("/manager/jobcan/holiday-types", methods=["GET"])
@token_required
@manager_required
def get_manager_jobcan_holiday_types():
    """
    管理画面「休暇タイプ一覧」用: Jobcan API から休暇タイプ一覧を取得する。
    scripts/test_jobcan_holiday_types.py と同様（APP_ENV に応じて sandbox / production）。
    """
    try:
        from services.jobcan_service import JobcanService

        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"

        jobcan_service = JobcanService(
            db=db,
            sandbox=is_sandbox,
            raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES,
        )
        result = jobcan_service.get_holiday_types()
        if result is None:
            abort(502, "Jobcan API から休暇タイプの取得に失敗しました。")

        # Firestore holiday_types コレクションへ上書き保存（doc ID = holiday_type_id）
        try:
            saved_count = save_jobcan_holiday_types_to_firestore(db, result)
        except Exception as save_err:
            current_app.logger.error(f"save_jobcan_holiday_types_to_firestore: {save_err}")
            abort(500, f"Firestore への保存に失敗しました: {save_err}")

        if isinstance(result, dict):
            response_body = dict(result)
        elif isinstance(result, list):
            response_body = {"holiday_types": result}
        else:
            response_body = {"data": result}
        # holiday.start/end から算出した minutes を各要素に付与（API 応答・Firestore と整合）
        enrich_holiday_types_payload_with_minutes(response_body)
        response_body["saved_to_firestore"] = saved_count
        return jsonify(response_body), 200
    except Exception as e:
        current_app.logger.error(f"get_manager_jobcan_holiday_types: {e}")
        abort(500, str(e))


def _normalize_jobcan_employee_id_for_match(value):
    """
    Jobcan master employees の id と users.jobcan_employee_id を突き合わせるための正規化。
    数値なら int 相当の文字列にそろえる（"102", 102, 102.0 を同一視）。
    """
    if value is None or isinstance(value, bool):
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


@api_bp.route("/manager/jobcan/employees", methods=["GET"])
@token_required
@manager_required
def get_manager_jobcan_employees_summary():
    """
    管理画面「スタッフ一覧」用: Jobcan GET /master/v1/employees を全件取得し、
    一覧表示用に id, last_name, first_name, main_group, sub_group, work_kind のみ返す。
    work_kind=4（非常勤役員）はレスポンスから除外する。

    併せて users コレクションを更新する:
    users.jobcan_employee_id が Jobcan の id と一致するドキュメントに、
    Jobcan の work_kind を work_kind_id として、entered_day を entered_day として書き込む。
    （work_kind は数値にできる場合は int）
    """
    try:
        from services.jobcan_service import JobcanService

        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"

        jobcan_service = JobcanService(
            db=db,
            sandbox=is_sandbox,
            raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES,
        )
        result = jobcan_service.get_all_employees()
        if result is None:
            abort(502, "Jobcan API から従業員一覧の取得に失敗しました。")

        raw_list = result.get("employees") or []

        # --- users: jobcan_employee_id 一致ドキュメントへ work_kind_id / entered_day を反映 ---
        users_work_kind_updated = 0
        try:
            jc_key_to_user = {}
            for udoc in db.collection("users").stream():
                udata = udoc.to_dict() or {}
                jkey = _normalize_jobcan_employee_id_for_match(udata.get("jobcan_employee_id"))
                if jkey:
                    jc_key_to_user[jkey] = {
                        "ref": udoc.reference,
                        "work_kind_id": udata.get("work_kind_id"),
                        "entered_day": udata.get("entered_day"),
                    }

            batch = db.batch()
            batch_ops = 0
            max_batch = 450

            for e in raw_list:
                if not isinstance(e, dict):
                    continue
                ekey = _normalize_jobcan_employee_id_for_match(e.get("id"))
                if not ekey:
                    continue
                user_obj = jc_key_to_user.get(ekey)
                if not user_obj:
                    continue

                wk_raw = e.get("work_kind")
                wk_val = None
                if wk_raw is not None:
                    try:
                        wk_val = int(wk_raw)
                    except (TypeError, ValueError):
                        wk_val = wk_raw

                entered_day_raw = e.get("entered_day")
                entered_day_val = None
                if entered_day_raw is not None:
                    entered_day_val = str(entered_day_raw).strip() or None

                update_payload = {}
                if user_obj.get("work_kind_id") != wk_val:
                    update_payload["work_kind_id"] = wk_val
                if user_obj.get("entered_day") != entered_day_val:
                    update_payload["entered_day"] = entered_day_val
                if not update_payload:
                    continue

                update_payload["updated_at"] = firestore.SERVER_TIMESTAMP
                batch.update(user_obj["ref"], update_payload)
                users_work_kind_updated += 1
                batch_ops += 1
                if batch_ops >= max_batch:
                    batch.commit()
                    batch = db.batch()
                    batch_ops = 0

            if batch_ops > 0:
                batch.commit()
        except Exception as sync_err:
            current_app.logger.error(f"sync work_kind_id to users failed: {sync_err}")
            abort(500, f"users の work_kind_id 更新に失敗しました: {sync_err}")

        slim = []
        for e in raw_list:
            if not isinstance(e, dict):
                continue
            # 非常勤役員（work_kind=4）は一覧に含めない
            wk = e.get("work_kind")
            if wk is not None:
                try:
                    if int(wk) == 4:
                        continue
                except (TypeError, ValueError):
                    pass
            sub = e.get("sub_group")
            if not isinstance(sub, list):
                sub = []
            slim.append(
                {
                    "id": e.get("id"),
                    "last_name": e.get("last_name"),
                    "first_name": e.get("first_name"),
                    "main_group": e.get("main_group"),
                    "sub_group": sub,
                    "work_kind": e.get("work_kind"),
                    "entered_day": e.get("entered_day"),
                }
            )

        return jsonify(
            {
                "employees": slim,
                "count": len(slim),
                "users_work_kind_updated": users_work_kind_updated,
            }
        ), 200
    except Exception as e:
        current_app.logger.error(f"get_manager_jobcan_employees_summary: {e}")
        abort(500, str(e))


@api_bp.route("/manager/categories/net/mapping", methods=["POST"])
@token_required
@manager_required
def save_net_category_mapping():
    """
    ネット事業部カテゴリ設定（マッピング）を保存する。
    リクエストボディは以下の形式:
    {
        "cat_b_id_1": {
            "cat_a_id_1": { "active": true, "color": "#FFFFFF", "label": "...", "sort": 10 },
            "cat_a_id_2": { "active": false, "color": "#000000", "label": "...", "sort": 20 }
        },
        ...
    }
    sort: PC入力画面での業務種別の並び順（数値・省略可）。category_a_sort に保存される。
    """
    mapping_data = request.get_json()
    if not isinstance(mapping_data, dict):
        abort(400, "Request body must be a JSON object.")

    try:
        batch = db.batch()
        cat_b_ref = db.collection("category_b")

        for cat_b_id, cat_a_mappings in mapping_data.items():
            if not isinstance(cat_a_mappings, dict):
                continue

            new_settings = {}
            new_sort = {}
            for cat_a_id, details in cat_a_mappings.items():
                # activeがtrueで、colorが設定されているものだけをマップに含める
                if details.get("active") and details.get("color") and str(details["color"]).startswith('#'):
                    new_settings[cat_a_id] = str(details["color"]).upper()
                # PC入力画面専用の並び順（sort が数値の場合のみ category_a_sort に含める）
                sort_val = details.get("sort")
                if sort_val is not None and isinstance(sort_val, (int, float)) and not isinstance(sort_val, bool):
                    new_sort[cat_a_id] = int(sort_val)

            doc_ref = cat_b_ref.document(cat_b_id)
            batch.update(doc_ref, {
                "category_a_settings": new_settings,
                "category_a_sort": new_sort
            })

        batch.commit()
        return jsonify({"status": "success", "message": f"{len(mapping_data)}件のカテゴリマッピングを更新しました。"}), 200
    except Exception as e:
        current_app.logger.error(f"Failed to save net category mapping: {e}", exc_info=True)
        abort(500, f"マッピングの保存中にエラーが発生しました: {e}")

@api_bp.route("/manager/categories/b/update_status", methods=["POST"])
@token_required
@manager_required
def update_category_b_status():
    """
    管理者向けに、複数のcategory_bのactiveステータスを一括更新する。
    """
    updates = request.get_json()
    if not isinstance(updates, list):
        abort(400, "Request body must be a list of category updates.")
    
    try:
        update_category_b_statuses(updates)
        return jsonify({"status": "success"}), 200
    except Exception as e:
        abort(500, f"An error occurred while updating categories: {e}")

@api_bp.route("/manager/groups", methods=["GET"])
@token_required
@login_required
def get_all_groups():
    """
    全てのグループマッピング情報を取得する。
    """
    try:
        docs = db.collection("group_mappings").stream()
        groups = []
        for doc in docs:
            data = doc.to_dict()
            groups.append({
                "id": doc.id,
                "name": data.get("name", "名称未設定")
            })
        # ID順でソート
        groups.sort(key=lambda x: int(x['id']) if x['id'].isdigit() else x['id'])
        return jsonify(groups), 200
    except Exception as e:
        current_app.logger.error(f"Failed to fetch groups: {e}")
        abort(500, "Failed to fetch groups.")

@api_bp.route("/manager/categories/b/check", methods=["GET"])
@token_required
@manager_required
def check_category_b_exists():
    """
    指定されたkindとlabelを持つカテゴリが存在するかチェックする。
    """
    kind = request.args.get("kind")
    label = request.args.get("label")
    if not kind or not label:
        abort(400, "Query parameters 'kind' and 'label' are required.")

    query = db.collection("category_b").where(filter=FieldFilter("label", "==", label)).where(filter=FieldFilter("kind", "==", kind)).limit(1)
    docs = list(query.stream())
    
    if not docs:
        return jsonify({"status": "not_exists"})

    doc_data = docs[0].to_dict()
    is_active = doc_data.get("active", False)

    if is_active:
        return jsonify({"status": "active"})
    else:
        return jsonify({"status": "inactive"})

@api_bp.route("/manager/categories/b/create", methods=["POST"])
@token_required
@manager_required
def create_category_b():
    """
    新しいcategory_bを作成する。
    """
    data = request.get_json()
    label = data.get("label")
    kind = data.get("kind")
    client = data.get("client")
    project = data.get("project")
    offices = data.get("offices")

    if not label or not kind:
        abort(400, "Request body must contain 'label' and 'kind'.")

    try:
        new_category = create_new_category_b(label=label, kind=kind, client=client, project=project, offices=offices)
        return jsonify(new_category), 201 # 201 Created
    except Exception as e:
        abort(500, f"An error occurred while creating the category: {e}")

@api_bp.route("/manager/categories/b/update", methods=["POST"])
@token_required
@manager_required
def update_category_b():
    """
    category_bの情報を更新する（顧客、案件、事業所）。
    """
    data = request.get_json()
    doc_id = data.get("id")
    client = data.get("client")
    project = data.get("project")
    offices = data.get("offices")

    if not doc_id:
        abort(400, "Request body must contain 'id'.")

    try:
        result = update_category_b_details(doc_id, client, project, offices)
        return jsonify(result), 200
    except Exception as e:
        current_app.logger.error(f"Failed to update category: {e}")
        abort(500, f"An error occurred while updating the category: {e}")

@api_bp.route("/manager/categories/b/update_offices", methods=["POST"])
@token_required
@manager_required
def update_category_b_offices_endpoint():
    """
    管理者向けに、複数のcategory_bのofficesを一括更新する。
    """
    updates = request.get_json()
    if not isinstance(updates, list):
        abort(400, "Request body must be a list of updates.")
    
    try:
        update_category_b_offices(updates)
        return jsonify({"status": "success"}), 200
    except Exception as e:
        current_app.logger.error(f"Failed to update offices: {e}")
        abort(500, f"An error occurred while updating offices: {e}")

@api_bp.route("/manager/categories/b/reactivate", methods=["POST"])
@token_required
@manager_required
def reactivate_category_b_endpoint():
    """
    非表示のカテゴリを再表示（active: trueに更新）する。
    """
    data = request.get_json()
    label = data.get("label")
    kind = data.get("kind")

    if not label or not kind:
        abort(400, "Request body must contain 'label' and 'kind'.")

    try:
        updated_category = reactivate_category_b(label=label, kind=kind)
        return jsonify(updated_category), 200
    except Exception as e:
        abort(500, f"An error occurred while reactivating the category: {e}")


@api_bp.route("/categories/b", methods=["GET"])
@token_required
def get_category_b_for_user():
    """
    一般ユーザー向けに、category_bのラベル一覧を降順で返すエンドポイント。
    工数入力画面の選択肢として使用される。
    """
    try:
        # 1. Firestoreから基本情報を取得
        user_info = get_user_info_by_line_id(g.line_user_id)
        jobcan_employee_id = user_info.get("jobcan_employee_id")
        current_app.logger.info(
            "[/api/categories/b] start line_user_id=%s company_employee_id=%s fs_main_group=%s fs_main_group_name=%s has_jobcan_employee_id=%s",
            g.line_user_id,
            user_info.get("company_employee_id"),
            user_info.get("main_group"),
            user_info.get("main_group_name"),
            bool(jobcan_employee_id),
        )

        # 2. Jobcan APIから最新のグループ情報を取得
        if jobcan_employee_id:
            from services.jobcan_service import JobcanService
            app_env = os.environ.get("APP_ENV", "development")
            is_sandbox = app_env != "production"
            jobcan_service = JobcanService(sandbox=is_sandbox)
            employee_details = jobcan_service.get_employee_details(employee_code=jobcan_employee_id)
            
            latest_group_id = employee_details.get("main_group") if employee_details else None

            if latest_group_id is not None:
                # Jobcanから取得したIDを元に、group_mappingsからグループ名を取得
                group_mapping_ref = db.collection("group_mappings").document(str(latest_group_id))
                group_mapping_doc = group_mapping_ref.get()
                
                if group_mapping_doc.exists:
                    latest_group_name = group_mapping_doc.to_dict().get("name")
                    # このリクエスト内で使用するグループ名を最新のものに更新
                    user_info["main_group_name"] = latest_group_name
                else:
                    # マッピングが見つからない場合、ログを出力してフォールバック
                    current_app.logger.warning(f"Group mapping not found for ID: {latest_group_id}. User: {g.line_user_id}")

                # Firestoreに保存されているグループIDを取得
                user_doc_ref = db.collection("users").document(user_info["company_employee_id"])
                original_user_doc = user_doc_ref.get()
                original_group_id = original_user_doc.to_dict().get("main_group") if original_user_doc.exists else None

                # FirestoreのIDとJobcanのIDが異なる場合、DBを更新
                if latest_group_id != original_group_id:
                    current_app.logger.info(f"Updating main_group ID for user {g.line_user_id} from '{original_group_id}' to '{latest_group_id}'.")
                    user_doc_ref.update({
                        "main_group": latest_group_id,
                        "updated_at": firestore.SERVER_TIMESTAMP
                    })
            else:
                current_app.logger.warning(f"Could not retrieve latest group info from Jobcan for user {g.line_user_id}. Falling back to Firestore data.")

        # 3. カテゴリ kind を決定
        #    優先順位:
        #    (1) Jobcan から取得した最新 main_group
        #    (2) users.main_group（Firestore）
        #    (3) main_group_name の文字列判定（後方互換フォールバック）
        resolved_main_group = None
        if jobcan_employee_id:
            # 上で取得できていれば latest_group_id がローカル変数に入っている想定
            try:
                if latest_group_id is not None:
                    resolved_main_group = str(latest_group_id).strip()
            except Exception:
                resolved_main_group = None

        if not resolved_main_group:
            fs_main_group = user_info.get("main_group")
            if fs_main_group is not None:
                resolved_main_group = str(fs_main_group).strip()

        if resolved_main_group == "3":
            kind = "net"
        else:
            current_group_name = user_info.get("main_group_name", "") or ""
            if "ネット" in current_group_name:
                kind = "net"
            else:
                # デフォルトは工務
                kind = "engineering"
        current_app.logger.info(
            "[/api/categories/b] resolved line_user_id=%s latest_group_id=%s resolved_main_group=%s resolved_group_name=%s kind=%s",
            g.line_user_id,
            locals().get("latest_group_id"),
            resolved_main_group,
            user_info.get("main_group_name", ""),
            kind,
        )

        # 4. 決定したkindでカテゴリを取得して返す (詳細情報付き)
        # get_all_category_b_labels(kind=kind) の代わりに直接クエリを実行
        query = db.collection("category_b").where(filter=FieldFilter("kind", "==", kind)).where(filter=FieldFilter("active", "==", True))
        docs = query.stream()

        categories = []
        for doc in docs:
            data = doc.to_dict()
            item = {
                "id": doc.id,
                "label": data.get("label", ""),
                "client": data.get("client", ""),
                "project": data.get("project", ""),
                "order": data.get("order", 0),
                "offices": data.get("offices", [])
            }
            # ネット用: 集計(B)に紐づく業務(A)のID一覧を返す（B先選択→A絞り込みに利用）
            if kind == "net":
                item["category_a_settings"] = data.get("category_a_settings", {})
            categories.append(item)

        # order降順、label昇順でソート
        categories.sort(key=lambda x: (-x['order'], x['label']))
        current_app.logger.info(
            "[/api/categories/b] result line_user_id=%s kind=%s category_count=%d",
            g.line_user_id,
            kind,
            len(categories),
        )
        if len(categories) == 0:
            current_app.logger.warning(
                "[/api/categories/b] empty categories line_user_id=%s kind=%s resolved_main_group=%s group_name=%s",
                g.line_user_id,
                kind,
                resolved_main_group,
                user_info.get("main_group_name", ""),
            )

        return jsonify(categories), 200
    except Exception as e:
        current_app.logger.error(f"Error getting categories for user {g.line_user_id}: {e}")
        # エラーが発生した場合は、デフォルトの工務カテゴリを返す
        labels = get_all_category_b_labels(kind="engineering")
        return jsonify(labels)

@api_bp.route("/system-notices", methods=["GET"])
@token_required
def get_system_notices():
    """
    ダッシュボード表示用のシステム改修予定・履歴を取得する。
    """
    try:
        # 日付の降順（新しい順）で取得
        docs = db.collection("system_notices").order_by("date", direction=firestore.Query.DESCENDING).stream()
        notices = []
        for doc in docs:
            data = doc.to_dict()
            notices.append({
                "id": doc.id,
                "type": data.get("type", "history"), # plan or history
                "date": data.get("date", ""),
                "content": data.get("content", "")
            })
        return jsonify(notices), 200
    except Exception as e:
        current_app.logger.error(f"Failed to fetch system notices: {e}")
        return jsonify([]), 200 # エラー時は空リストを返して画面を壊さない

@api_bp.route("/system-notices", methods=["POST"])
@token_required
@manager_required
def create_system_notice():
    """システム改修情報（お知らせ）を新規作成する。"""
    data = request.get_json()
    if not data or not data.get('date') or not data.get('content'):
        abort(400, "dateとcontentは必須です。")

    try:
        new_notice = {
            "type": data.get("type", "history"),
            "date": data.get("date"),
            "content": data.get("content"),
            "created_at": firestore.SERVER_TIMESTAMP
        }
        # add() は (timestamp, DocumentReference) のタプルを返す
        _, doc_ref = db.collection("system_notices").add(new_notice)
        
        created_data = new_notice.copy()
        created_data['id'] = doc_ref.id
        created_data.pop('created_at', None) # TimestampはJSONシリアライズ不可
        
        return jsonify(created_data), 201
    except Exception as e:
        current_app.logger.error(f"Failed to create system notice: {e}")
        abort(500, "お知らせの作成に失敗しました。")

@api_bp.route("/system-notices/<notice_id>", methods=["PUT"])
@token_required
@manager_required
def update_system_notice(notice_id):
    """システム改修情報（お知らせ）を更新する。"""
    data = request.get_json()
    if not data or not data.get('date') or not data.get('content'):
        abort(400, "dateとcontentは必須です。")

    try:
        doc_ref = db.collection("system_notices").document(notice_id)
        update_data = {
            "type": data.get("type"),
            "date": data.get("date"),
            "content": data.get("content"),
            "updated_at": firestore.SERVER_TIMESTAMP
        }
        doc_ref.update(update_data)
        return jsonify({"status": "success", "id": notice_id}), 200
    except Exception as e:
        current_app.logger.error(f"Failed to update system notice {notice_id}: {e}")
        abort(500, "お知らせの更新に失敗しました。")

@api_bp.route("/system-notices/<notice_id>", methods=["DELETE"])
@token_required
@manager_required
def delete_system_notice(notice_id):
    """システム改修情報（お知らせ）を削除する。"""
    try:
        db.collection("system_notices").document(notice_id).delete()
        return jsonify({"status": "success", "message": "削除しました。"}), 200
    except Exception as e:
        current_app.logger.error(f"Failed to delete system notice {notice_id}: {e}")
        abort(500, "お知らせの削除に失敗しました。")

def _run_batch_sync_today_work_times_for_all_employees():
    """
    Cloud Scheduler用:
    全アクティブ従業員の「当日」勤務時間を Jobcan 側で再集計してから取得し、
    Firestore daily_reports.jobcan_work_minutes を更新する。
    """
    try:
        # 1. 対象日付（当日）
        jst = timezone(timedelta(hours=9))
        target_date_str = datetime.now(jst).strftime('%Y-%m-%d')
        current_app.logger.info(f"Batch job started: Sync all employees work times for {target_date_str}")

        # 2. 全アクティブ従業員を取得
        mappings_ref = db.collection('employee_mappings')
        query = mappings_ref.where(filter=FieldFilter("status", "==", "active"))
        mappings_docs = query.stream()

        employees = []
        for doc in mappings_docs:
            data = doc.to_dict()
            if data.get('jobcan_employee_id'):
                employees.append({
                    'company_employee_id': doc.id,
                    'jobcan_employee_id': data['jobcan_employee_id']
                })

        if not employees:
            return jsonify({"status": "success", "message": "No active employees found."}), 200

        # 3. JobcanService初期化
        from services.jobcan_service import JobcanService
        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"
        jobcan_service = JobcanService(
            db=db,
            sandbox=is_sandbox,
            raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES
        )

        updated_count = 0
        skipped_no_report_count = 0
        error_count = 0

        # 4. 逐次処理で更新
        for emp in employees:
            try:
                # 当日は refresh API -> 待機 -> summaries API の順で強制同期
                work_minutes = jobcan_service.calculate_work_time_from_adits(
                    employee_id=emp['jobcan_employee_id'], 
                    date=target_date_str, 
                    wait_seconds=2.0
                )
                
                # Firestore更新
                doc_id = f"{emp['company_employee_id']}_{target_date_str}"
                doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
                # ドキュメントが存在する場合のみ更新（未提出の場合は更新しない）
                # ※必要であれば set(..., merge=True) にしてドキュメントを作成することも可能
                try:
                    doc_ref.update({"jobcan_work_minutes": work_minutes})
                    updated_count += 1
                except Exception:
                    # 日報未作成の場合はスキップ（またはログ出力）
                    skipped_no_report_count += 1

            except Exception as e:
                current_app.logger.error(f"Failed to refresh work time for {emp['company_employee_id']}: {e}")
                error_count += 1

        return jsonify({
            "status": "success",
            "processed": len(employees),
            "updated": updated_count,
            "skipped_no_report": skipped_no_report_count,
            "errors": error_count
        }), 200

    except Exception as e:
        current_app.logger.error(f"Batch job failed: {e}")
        abort(500, f"Batch job failed: {e}")


@api_bp.route("/batch/sync-jobcan-today-all-employees", methods=["POST"])
@scheduler_token_required
def batch_sync_jobcan_today_all_employees():
    """
    Cloud Scheduler専用:
    当日分 + 全アクティブ従業員の勤務時間を Jobcan と一斉同期する。
    """
    return _run_batch_sync_today_work_times_for_all_employees()


@api_bp.route("/batch/refresh-all-work-times", methods=["POST"])
@scheduler_token_required
def batch_refresh_all_work_times():
    """
    旧エンドポイント（互換維持）。
    内部的には /batch/sync-jobcan-today-all-employees と同じ処理を実行する。
    """
    return _run_batch_sync_today_work_times_for_all_employees()

@api_bp.route("/batch/notify-unreported", methods=["POST"])
@scheduler_token_required
def batch_notify_unreported():
    """
    Cloud Schedulerからトリガーされ、工数未入力のユーザーに通知を送信する。
    """
    data = request.get_json()
    notify_time_str = data.get("notify_time", "16:00") # "15:30" などの形式
    if not notify_time_str:
        abort(400, "Request body must contain 'notify_time'.")

    today_str = (datetime.now(timezone(timedelta(hours=9)))).strftime('%Y-%m-%d')
    current_app.logger.info(f"Batch job started for notify_time: {notify_time_str} on {today_str}")

    # 1. 通知対象時刻のユーザーを取得
    users_ref = db.collection("users")
    # `notify_time` フィールドはFirestoreに文字列 "HH:MM" で保存されている想定
    query = users_ref.where(filter=FieldFilter("notify_time", "==", notify_time_str))
    target_users = list(query.stream())

    if not target_users:
        current_app.logger.info("No users found for this notification time.")
        return jsonify({"status": "success", "message": "No users to notify."}), 200

    # 2. JobcanServiceを初期化
    from services.jobcan_service import JobcanService
    app_env = os.environ.get("APP_ENV", "development")
    is_sandbox = app_env != "production"
    jobcan_service = JobcanService(
        db=db,
        sandbox=is_sandbox,
        raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES
    )

    notified_count = 0
    # 3. 各ユーザーの条件をチェックして通知
    for user_doc in target_users:
        user_data = user_doc.to_dict()
        line_user_id = user_data.get("line_user_id")
        jobcan_employee_id = user_data.get("jobcan_employee_id")
        company_employee_id = user_doc.id

        if not all([line_user_id, jobcan_employee_id, company_employee_id]):
            current_app.logger.warning(f"Skipping user {user_doc.id} due to missing info.")
            continue

        try:
            # 条件1: 当日の勤務時間を取得
            summary_res = jobcan_service.get_daily_summaries(employee_id=jobcan_employee_id, dates=[today_str])
            work_minutes = 0
            if summary_res and summary_res.get("daily_summaries"):
                work_minutes = summary_res["daily_summaries"][0].get("work", 0)

            # 条件2: 当日の工数報告時間を取得
            report_doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(f"{company_employee_id}_{today_str}")
            report_doc = report_doc_ref.get()
            reported_minutes = report_doc.to_dict().get("task_total_minutes", 0) if report_doc.exists else 0

            # 条件判定と通知実行
            if work_minutes > 0 and reported_minutes == 0:
                current_app.logger.info(f"Notifying user {company_employee_id} (LINE: {line_user_id}). Work: {work_minutes}min, Reported: {reported_minutes}min.")
                send_push_message(line_user_id, "本日の工数が未入力です。")
                notified_count += 1
            else:
                current_app.logger.info(f"Skipping user {company_employee_id}. Work: {work_minutes}min, Reported: {reported_minutes}min.")

        except Exception as e:
            current_app.logger.error(f"Error processing user {company_employee_id}: {e}")

    return jsonify({
        "status": "success",
        "processed_users": len(target_users),
        "notified_users": notified_count
    }), 200

@api_bp.route("/manager/employees/check-unmapped", methods=["GET"])
@token_required
def get_unmapped_employees():
    """
    Jobcanに存在するが、Firestoreにマッピングされていない従業員のリストと、
    Jobcanの全従業員数を返す。
    """
    try:
        result = check_unmapped_jobcan_employees()
        return jsonify(result), 200
    except Exception as e:
        abort(500, f"未マッピング従業員のチェック中にエラーが発生しました: {e}")

@api_bp.route("/manager/employees/create-mapping", methods=["POST"])
@token_required
def post_employee_mapping():
    """
    新しい従業員マッピングをFirestoreに作成する。
    """
    data = request.get_json()
    company_employee_id = data.get("company_employee_id")
    jobcan_employee_id = data.get("jobcan_employee_id")
    name = data.get("name")

    if not all([company_employee_id, jobcan_employee_id, name]):
        abort(400, "company_employee_id, jobcan_employee_id, name は必須です。")

    try:
        new_mapping = create_employee_mapping(company_employee_id, jobcan_employee_id, name)
        return jsonify(new_mapping), 201 # 201 Created
    except Exception as e:
        # create_employee_mapping内でabortが呼ばれた場合、そのエラーが返される
        if hasattr(e, 'code'):
            raise
        abort(500, f"従業員マッピングの作成中にエラーが発生しました: {e}")

@api_bp.route("/categories/category_a", methods=["GET"])
@token_required
def get_category_a():
    """
    category_a（業務種別）のリストを取得するエンドポイント。
    工数入力画面の選択肢として使用される。
    """
    try:
        # クエリパラメータからkindを取得。なければ'engineering'をデフォルトとする
        kind = request.args.get("kind", "engineering")
        
        # 'engineering' または 'net' 以外が指定された場合も 'engineering' にフォールバック
        if kind not in ["engineering", "net"]:
            current_app.logger.warning(f"Invalid 'kind' parameter received: '{kind}'. Defaulting to 'engineering'.")
            kind = "engineering"

        current_app.logger.info(f"[/api/categories/category_a] Filtering by kind: '{kind}'")

        # クエリを構築。kindによるフィルタリングを常に行う
        query = db.collection("category_a").where(filter=FieldFilter("kind", "==", kind))

        # ★ orderでソートするクエリに戻します。
        # このクエリにはFirestoreの複合インデックスが必要になります。
        docs = query.order_by("order").stream()

        categories = []
        for doc in docs:
            data = doc.to_dict()
            if data.get("active", True) and data.get("label"):
                categories.append({"id": doc.id, "label": data.get("label")})
        return jsonify(categories), 200
    except Exception as e:
        current_app.logger.error(f"Error getting categories from 'category_a': {e}", exc_info=True)
        abort(500, "Failed to get categories from 'category_a'.")

@api_bp.route("/manager/categories/a", methods=["GET"])
@token_required
@login_required
def get_manager_category_a():
    """管理者向けに、category_aの一覧（詳細情報含む）を返すエンドポイント。"""
    try:        
        kind = request.args.get("kind")

        query = db.collection("category_a")

        # kindが指定されていれば、それでフィルタリング
        if kind:
            if kind not in ["engineering", "net"]:
                abort(400, "Invalid 'kind' parameter. Must be 'engineering' or 'net'.")
            query = query.where(filter=FieldFilter("kind", "==", kind))

        docs = query.stream()
        categories = []
        for doc in docs:
            data = doc.to_dict()
            # activeが明示的にfalseに設定されているドキュメントは除外
            if data.get("active") is False:
                continue
            categories.append({
                "id": doc.id,
                "label": data.get("label", ""),
                "active": data.get("active", True),
                "work_type": data.get("work_type"), # '加工', '現場', or None
                "order": data.get("order", 9999) # orderがない場合は大きな値
            })
        
        # Python側でorderキーを元にソート
        categories.sort(key=lambda x: x['order'])

        return jsonify(categories), 200
    except Exception as e:
        current_app.logger.error(f"Failed to fetch category_a list: {e}")
        abort(500, "業務種別一覧の取得に失敗しました。")

@api_bp.route("/manager/categories/a/batch-update", methods=["POST"])
@token_required
@manager_required
def batch_update_manager_category_a():
    """管理者向けに、category_aのリストを一括で更新する（順序とwork_type）。"""
    updates = request.get_json()
    if not isinstance(updates, list):
        abort(400, "Request body must be a list of category updates.")

    try:
        batch = db.batch()
        for item in updates:
            category_id = item.get("id")
            order = item.get("order")
            work_type = item.get("work_type")

            if not category_id or order is None: continue
            if work_type not in ['加工', '現場', None]: continue

            doc_ref = db.collection("category_a").document(category_id)
            batch.update(doc_ref, {"order": order, "work_type": work_type, "updated_at": firestore.SERVER_TIMESTAMP})
        
        batch.commit()
        return jsonify({"status": "success", "message": f"{len(updates)}件の業務種別を更新しました。"}), 200
    except Exception as e:
        current_app.logger.error(f"Failed to batch update category_a: {e}")
        abort(500, "業務種別の一括更新に失敗しました。")

@api_bp.route("/config", methods=["GET"])
@token_required
def get_config():
    """
    フロントエンドに提供する設定情報を返すエンドポイント。
    現在は締め日（CLOSING_DAY）を返す。
    """
    closing_day = int(os.environ.get("CLOSING_DAY", "20")) # 環境変数から取得、デフォルトは20
    return jsonify({"closing_day": closing_day}), 200

# --- 【実装保留】速報値メール送信機能 ---
# 代替案（GoogleドライブへのCSV保存）を実装するため、一旦コメントアウト。
# @api_bp.route("/manager/send-report-mail", methods=["POST"])
# @token_required
# @manager_required
# def send_report_mail():
#     """
#     管理者向けに、速報値メールを送信するエンドポイント。
#     メールは操作した管理者本人にのみ送信される。
#     """
#     try:
#         # @manager_required デコレータによって g.user_info にユーザー情報が格納されている
#         recipient_email = g.user_info.get("mail")
#
#         if not recipient_email:
#             abort(400, "あなたのユーザー情報にメールアドレスが登録されていません。")
#
#         send_quick_report_email(recipient_email=recipient_email)
#         return jsonify({"message": f"{recipient_email} に速報値メールを送信しました。"}), 200
#     except Exception as e:
#         # send_quick_report_email内でabortが呼ばれた場合、そのエラーが返される
#         # ここでは予期せぬエラーをキャッチする
#         abort(500, "メール送信処理の呼び出し中に予期せぬエラーが発生しました。")

@api_bp.route("/manager/daily-reports", methods=["GET"])
@token_required
@login_required
def get_manager_daily_reports():
    """
    指定された日付の日報一覧（予実突合データ）を取得する
    """
    target_date_str = request.args.get('date')
    if not target_date_str:
        abort(400, "Query parameter 'date' is required.")

    try:
        # 日付の形式チェック
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        
        # 1. 全社員マッピングを取得 (IDと名前のマッピング用)
        # これにより、未提出者も一覧に表示できる
        mappings_ref = db.collection('employee_mappings')
        # 日報_一覧では status: "active" に加え "active_officer"（日報を書く役員）も表示対象とする
        query = mappings_ref.where(filter=FieldFilter("status", "in", ["active", "active_officer"]))
        mappings_docs = query.stream()
        
        # usersコレクションからグループ情報を取得するためのマップを作成
        # employee_mappingsにはmain_groupがないため、usersから補完する
        users_ref = db.collection('users')
        users_docs = users_ref.stream()
        users_group_map = {}
        for u_doc in users_docs:
            u_data = u_doc.to_dict()
            # usersドキュメントIDはcompany_employee_idと同じ前提
            if 'main_group' in u_data:
                users_group_map[u_doc.id] = u_data['main_group']

        employees = []
        for doc in mappings_docs:
            data = doc.to_dict()
            emp_id = doc.id
            employees.append({
                'id': emp_id, # company_employee_id
                'name': data.get('name', 'Unknown'),
                'group_id': users_group_map.get(emp_id) # usersコレクションから取得したグループIDを設定
            })

        # 2. 指定日の日報データを取得
        reports_ref = db.collection(COLLECTION_DAILY_REPORTS)
        query = reports_ref.where(filter=FieldFilter("date", "==", target_date))
        reports_docs = query.stream()

        reports_map = {}
        for doc in reports_docs:
            data = doc.to_dict()
            emp_id = data.get('company_employee_id')
            if emp_id is not None:
                # キーは文字列に統一（doc.id と型を揃え、active_officer 等の日報も確実に突き合う）
                reports_map[str(emp_id)] = data

        # 3. 結合と集計
        results = []
        for emp in employees:
            emp_id = emp['id']
            name = emp['name']
            group_id = emp.get('group_id') # グループIDを取得
            report = reports_map.get(str(emp_id))
            
            if report:
                work_time = report.get('jobcan_work_minutes', 0)
                task_time = report.get('task_total_minutes', 0)
                diff = task_time - work_time
                
                results.append({
                    'date': target_date_str,
                    'name': name,
                    'workTime': work_time,
                    'taskTime': task_time,
                    'diff': diff,
                    'employeeId': emp_id,
                    'group_id': group_id # レスポンスに含める
                })
            else:
                # 未提出者
                results.append({
                    'date': target_date_str,
                    'name': name,
                    'workTime': None,
                    'taskTime': None,
                    'diff': None,
                    'employeeId': emp_id,
                    'group_id': group_id # レスポンスに含める
                })

        # 社員ID順でソート
        results.sort(key=lambda x: x['employeeId'])

        return jsonify(results), 200

    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")
    except Exception as e:
        current_app.logger.error(f"Error fetching daily reports: {e}")
        abort(500, f"Internal Server Error: {str(e)}")

@api_bp.route("/manager/user-by-employee-id", methods=["GET"])
@token_required
@login_required
def get_manager_user_by_employee_id():
    """
    社員IDから従業員情報を取得する（管理者用）。
    日報_スタッフ別画面の検索機能で使用。
    """
    target_id = request.args.get("employee_id")
    if not target_id:
        abort(400, "Query parameter 'employee_id' is required.")

    # employee_mappings から基本情報を取得
    mapping_ref = db.collection("employee_mappings").document(target_id)
    mapping_doc = mapping_ref.get()

    if not mapping_doc.exists:
        abort(404, "該当する社員IDが見つかりません。")

    mapping_data = mapping_doc.to_dict()
    
    # users コレクションからグループ情報を取得 (存在する場合)
    user_ref = db.collection("users").document(target_id)
    user_doc = user_ref.get()
    group_id = None
    is_executive = False
    if user_doc.exists:
        user_data = user_doc.to_dict()
        group_id = user_data.get("main_group")
        is_executive = user_data.get("is_executive", False)

    # 履歴情報の取得
    history = get_user_selection_history(target_id)

    return jsonify({
        "employeeId": target_id,
        "name": mapping_data.get("name", "名称未設定"),
        "groupId": group_id,
        "is_executive": is_executive,
        "history": history
    }), 200

@api_bp.route("/manager/calendar-statuses", methods=["GET"])
@token_required
@login_required
def get_manager_calendar_statuses():
    """
    指定された従業員・期間のカレンダーステータスを取得する（管理者用）。
    LIFF用の get_calendar_statuses ユーティリティを再利用する。
    """
    target_employee_id = request.args.get("employee_id")
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    if not all([target_employee_id, start_date_str, end_date_str]):
        abort(400, "Missing required parameters: employee_id, start_date, end_date")

    # 対象従業員の Jobcan ID を取得
    mapping_ref = db.collection("employee_mappings").document(target_employee_id)
    mapping_doc = mapping_ref.get()
    
    if not mapping_doc.exists:
        abort(404, "Target user not found.")
        
    jobcan_employee_id = mapping_doc.to_dict().get("jobcan_employee_id")
    
    # 日付リスト生成
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        if start_date > end_date:
            abort(400, "start_date must be before end_date")
            
        dates_in_month = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end_date - start_date).days + 1)]
    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")

    # 既存の共通ロジック (app_core/utils.py) を再利用してステータスを生成
    statuses = get_calendar_statuses(jobcan_employee_id, target_employee_id, dates_in_month)
    
    return jsonify(statuses), 200

@api_bp.route("/manager/past-reports", methods=["GET"])
@token_required
@login_required
def get_past_reports():
    """
    指定された従業員・期間の過去の日報データを取得する。
    ネット事業部の日報_個別・代理入力など、登録ユーザー全員が参照できるようにする。
    """
    employee_id = request.args.get('employee_id')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not all([employee_id, start_date_str, end_date_str]):
        abort(400, "employee_id, start_date, end_date are required.")

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")

    try:
        return jsonify(_fetch_past_reports_by_date(employee_id, start_date, end_date)), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching past reports for {employee_id}: {e}", exc_info=True)
        abort(500, "過去日報の取得中にエラーが発生しました。")


@api_bp.route("/past-reports", methods=["GET"])
@token_required
@login_required
def get_past_reports_for_user():
    """
    一般ユーザー向け過去日報取得。
    employee_id 未指定時は本人。指定時も登録ユーザーであれば任意の従業員を参照可能（権限による制限なし）。
    """
    user_info = get_user_info_by_line_id(g.line_user_id)
    self_employee_id = user_info.get("company_employee_id")

    employee_id = request.args.get('employee_id') or self_employee_id
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not all([employee_id, start_date_str, end_date_str]):
        abort(400, "employee_id, start_date, end_date are required.")

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")

    try:
        return jsonify(_fetch_past_reports_by_date(employee_id, start_date, end_date)), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching user past reports for {employee_id}: {e}", exc_info=True)
        abort(500, "過去日報の取得中にエラーが発生しました。")


def _fetch_past_reports_by_date(employee_id: str, start_date: datetime, end_date: datetime) -> dict:
    query = db.collection(COLLECTION_DAILY_REPORTS) \
        .where(filter=FieldFilter("company_employee_id", "==", employee_id)) \
        .where(filter=FieldFilter("date", ">=", start_date)) \
        .where(filter=FieldFilter("date", "<=", end_date)) \
        .order_by("date")

    docs = query.stream()
    reports_by_date = {}
    for doc in docs:
        data = doc.to_dict()
        report_date = data.get('date')
        if isinstance(report_date, datetime):
            date_key = report_date.strftime('%Y-%m-%d')
            raw_tasks = data.get('tasks', []) or []
            normalized_tasks = []
            for task in raw_tasks:
                if not isinstance(task, dict):
                    continue
                task_copy = dict(task)
                if 'comment' not in task_copy:
                    task_copy['comment'] = ""
                normalized_tasks.append(task_copy)

            reports_by_date[date_key] = normalized_tasks

    return reports_by_date

@api_bp.route("/manager/work-times", methods=["GET"])
@token_required
@login_required
def get_all_employee_work_times():
    """
    【管理者向け】指定された日付の全アクティブ従業員の勤務時間をJobcanから取得する。
    日報＞グループ別画面の「勤務時間受信」ボタンで使用される。
    """
    target_date_str = request.args.get('date')
    if not target_date_str:
        abort(400, "Query parameter 'date' is required.")

    try:
        # 日付の形式チェック
        datetime.strptime(target_date_str, '%Y-%m-%d')
    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")

    try:
        # 1. 全アクティブ従業員のマッピングを取得 (jobcan_employee_id を含む)
        mappings_ref = db.collection('employee_mappings')
        query = mappings_ref.where(filter=FieldFilter("status", "==", "active"))
        mappings_docs = query.stream()

        employees_to_fetch = []
        for doc in mappings_docs:
            data = doc.to_dict()
            if data.get('jobcan_employee_id'):
                employees_to_fetch.append({
                    'company_employee_id': doc.id,
                    'jobcan_employee_id': data['jobcan_employee_id']
                })

        if not employees_to_fetch:
            return jsonify([]), 200

        # 2. JobcanServiceを初期化
        from services.jobcan_service import JobcanService
        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"
        jobcan_service = JobcanService(sandbox=is_sandbox)

        # 3. 各従業員の勤務時間を取得
        results = []
        for emp in employees_to_fetch:
            company_id = emp['company_employee_id']
            jobcan_id = emp['jobcan_employee_id']
            try:
                response_data = jobcan_service.get_daily_summaries(employee_id=jobcan_id, dates=[target_date_str])
                work_minutes = 0
                if response_data and response_data.get("daily_summaries"):
                    work_minutes = response_data["daily_summaries"][0].get("work", 0)
                results.append({'employeeId': company_id, 'workTime': work_minutes})
            except Exception as e:
                current_app.logger.error(f"Failed to get work time for employee {jobcan_id} on {target_date_str}: {e}")
                results.append({'employeeId': company_id, 'workTime': None}) # エラー時はnullを返す
        return jsonify(results), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching all employee work times: {e}")
        abort(500, f"Internal Server Error: {str(e)}")

@api_bp.route("/manager/project-summary", methods=["GET"])
@token_required
@login_required
def get_project_summary():
    """
    指定された工事番号（プロジェクト）について、当月度の工数を集計し、
    日付を行、従業員を列とするマトリクス形式で返す。
    """
    # 1. リクエストパラメータの取得
    project_label = request.args.get('project_label')
    if not project_label:
        abort(400, "Query parameter 'project_label' is required.")

    # 基準日（省略時は今日）
    date_str = request.args.get('date', datetime.now(timezone(timedelta(hours=9))).strftime('%Y-%m-%d'))
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        abort(400, "Invalid date format for 'date'. Use YYYY-MM-DD.")

    # カテゴリ詳細（顧客名、案件名）を取得
    client = ""
    project = ""
    try:
        cat_query = db.collection("category_b").where(filter=FieldFilter("label", "==", project_label)).limit(1)
        cat_docs = list(cat_query.stream())
        if cat_docs:
            cat_data = cat_docs[0].to_dict()
            client = cat_data.get("client") or ""
            project = cat_data.get("project") or ""
    except Exception as e:
        current_app.logger.warning(f"Failed to fetch category details for {project_label}: {e}")

    try:
        # 2. 当月度の期間を計算
        start_date, end_date = calculate_monthly_period(target_date)

        # 2.5. 全てのアクティブな従業員情報を取得（列として使用）
        mappings_ref = db.collection('employee_mappings')
        query_mappings = mappings_ref.where(filter=FieldFilter("status", "==", "active"))
        mappings_docs = query_mappings.stream()
        
        users_ref = db.collection('users')
        users_docs = users_ref.stream()
        users_group_map = {doc.id: doc.to_dict().get('main_group') for doc in users_docs if doc.to_dict().get('main_group') is not None}

        all_active_employees = []
        for doc in mappings_docs:
            data = doc.to_dict()
            emp_id = doc.id
            if data.get('name'):
                all_active_employees.append({
                    'id': emp_id,
                    'name': data.get('name'),
                    'group_id': users_group_map.get(emp_id)
                })
        all_active_employees.sort(key=lambda x: x['name'])

        # 3. Firestoreから期間内のレポートを取得
        query = db.collection(COLLECTION_DAILY_REPORTS).where(filter=FieldFilter("date", ">=", start_date)).where(filter=FieldFilter("date", "<=", end_date))
        docs = query.stream()

        # 4. 指定された工事番号でフィルタリングし、データを集計
        # 形式: { "YYYY-MM-DD": { "employee_name": total_hours } }
        summary_data = {}
        relevant_employee_names = set() # 実績のある従業員名を保持するセット

        for doc in docs:
            data = doc.to_dict()
            report_date, employee_name, tasks = data.get('date'), data.get('employee_name'), data.get('tasks', [])
            if not all([report_date, employee_name]): continue

            project_minutes = sum(task.get('time', 0) for task in tasks if task.get('categoryB_label') == project_label)
            
            if project_minutes > 0:
                date_key = report_date.strftime('%Y-%m-%d')
                summary_data.setdefault(date_key, {})[employee_name] = summary_data.get(date_key, {}).get(employee_name, 0.0) + (project_minutes / 60.0)
                relevant_employee_names.add(employee_name)

        # 実績のある従業員のみにフィルタリング (all_active_employeesのソート順を維持)
        filtered_employees = [emp for emp in all_active_employees if emp['name'] in relevant_employee_names]

        # 5. マトリクス形式に整形
        rows = []
        current_date = start_date
        while current_date <= end_date:
            date_key = current_date.strftime('%Y-%m-%d')
            row = [date_key] + [summary_data.get(date_key, {}).get(emp['name'], 0.0) for emp in filtered_employees]
            rows.append(row)
            current_date += timedelta(days=1)

        return jsonify({
            "project_label": project_label, "client": client, "project": project,
            "start_date": start_date.strftime('%Y-%m-%d'), "end_date": end_date.strftime('%Y-%m-%d'),
            "employees": filtered_employees, "rows": rows
        }), 200

    except Exception as e:
        current_app.logger.error(f"Error fetching project summary for '{project_label}': {e}")
        abort(500, f"Internal Server Error: {str(e)}")

@api_bp.route("/manager/project-summary/excel", methods=["POST"])
@token_required
@login_required
def download_project_summary_excel():
    """
    工番別集計表のExcelファイルを生成して返す。
    当月度/前月度の実績から工事番号を動的に抽出し、シートを作成する。
    """
    try:
        data = request.get_json() or {}
        target_month = data.get('target_month', 'current') # 'current' or 'previous'

        # --- 1. 集計期間の決定 ---
        jst = timezone(timedelta(hours=9))
        base_date = datetime.now(jst)
        
        # まず基準日（今日）の月度範囲を取得
        start_date, end_date = calculate_monthly_period(base_date)

        if target_month == 'previous':
            # 当月度の開始日の前日を基準にして、前月度の範囲を再取得
            # これにより、締め日がいつであっても正確に前月度を取得できる
            prev_month_base = start_date - timedelta(days=1)
            start_date, end_date = calculate_monthly_period(prev_month_base)

        current_app.logger.info(f"Excel generation period for '{target_month}': {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        # --- categoryA_id に基づく工数分類の定義 ---
        KAKO_IDS = set()
        GENBA_IDS = set()
        YUKYU_IDS = set()
        try:
            cat_a_docs = db.collection("category_a").stream()
            for doc in cat_a_docs:
                data = doc.to_dict()
                w_type = data.get("work_type")
                if w_type == '加工':
                    KAKO_IDS.add(doc.id)
                elif w_type == '現場':
                    GENBA_IDS.add(doc.id)
                elif w_type == '有休':
                    YUKYU_IDS.add(doc.id)
        except Exception as e:
            current_app.logger.error(f"Failed to load category_a work_types: {e}")

        # --- 2. 期間内の日報データを取得し、集計 ---
        source_collection = _resolve_summary_source_collection(target_month, start_date, end_date, "enj")
        current_app.logger.info(
            f"project-summary/excel source collection for '{target_month}': {source_collection}"
        )
        query = db.collection(source_collection).where(filter=FieldFilter("date", ">=", start_date)).where(filter=FieldFilter("date", "<=", end_date))
        docs = list(query.stream()) # リスト化して保持

        unique_kouban_ids = set()
        # { "kouban_label": { "employee_name": { "YYYY-MM-DD": {"加工": minutes, "現場": minutes} } } }
        kouban_data_map = {}

        for doc in docs:
            data = doc.to_dict()
            report_date = data.get('date')
            employee_name = data.get('employee_name')
            
            if not report_date or not employee_name:
                continue
            
            date_str = report_date.strftime('%Y-%m-%d')
            tasks = data.get('tasks', [])
            
            for task in tasks:
                category_b_id = task.get('categoryB_id')
                if category_b_id and category_b_id.startswith('e_'):
                    unique_kouban_ids.add(category_b_id)
                    kouban_label = category_b_id.replace('e_', '')
                    category_a_id = task.get('categoryA_id')
                    time_minutes = task.get('time', 0)
                    
                    if kouban_label not in kouban_data_map:
                        kouban_data_map[kouban_label] = {}
                    if employee_name not in kouban_data_map[kouban_label]:
                        kouban_data_map[kouban_label][employee_name] = {}
                    if date_str not in kouban_data_map[kouban_label][employee_name]:
                        kouban_data_map[kouban_label][employee_name][date_str] = {'加工': 0, '現場': 0}

                    # 有休判定の強化: 工事番号が'000000'、またはカテゴリ名が'有休'、またはIDが'A00'/'A99'(有休・有休忌引)の場合も含める
                    is_yukyu = (
                        (kouban_label == '000000')
                        or (category_a_id in YUKYU_IDS)
                        or (task.get('categoryA_label') == '有休')
                        or (task.get('categoryA_label') == '有休(忌引)')
                        or (category_a_id == 'A00')
                        or (category_a_id == 'A99')
                    )

                    if is_yukyu:
                        # 有休は「現場」枠として集計し、シート生成時にヘッダー等を調整する
                        kouban_data_map[kouban_label][employee_name][date_str]['現場'] += time_minutes
                    elif category_a_id in KAKO_IDS:
                        kouban_data_map[kouban_label][employee_name][date_str]['加工'] += time_minutes
                    elif category_a_id in GENBA_IDS:
                        kouban_data_map[kouban_label][employee_name][date_str]['現場'] += time_minutes
        
        if not unique_kouban_ids:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "データなし"
            ws['A1'] = "指定された期間に工数入力のあった工事番号はありません。"
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            excel_data = output.read()
            b64_data = base64.b64encode(excel_data).decode('utf-8')
            file_name = f"工番別集計_{target_month}_no_data.xlsx"
            return jsonify({"file_name": file_name, "file_content": b64_data}), 200

        # 'e_240001' -> '240001'
        kouban_labels = sorted([kouban_id.replace('e_', '') for kouban_id in unique_kouban_ids], reverse=True)

        # --- 3. 工事番号に対応する顧客名・案件名を取得 ---
        category_info_map = {}
        chunk_size = 30
        for i in range(0, len(kouban_labels), chunk_size):
            chunk = kouban_labels[i:i + chunk_size]
            cat_query = db.collection("category_b").where(filter=FieldFilter("label", "in", chunk))
            cat_docs = cat_query.stream()
            for cat_doc in cat_docs:
                cat_data = cat_doc.to_dict()
                kouban_label = cat_data.get('label')
                if kouban_label:
                    category_info_map[kouban_label] = {
                        "client": cat_data.get("client", ""),
                        "project": cat_data.get("project", "")
                    }

        # --- 4. テンプレートを読み込み、シートを生成 ---
        template_path = os.path.join(current_app.root_path, 'temp', 'template.xlsx')
        if not os.path.exists(template_path):
            abort(500, f"テンプレートファイルが見つかりません: {template_path}")

        wb = openpyxl.load_workbook(template_path)
        try:
            template_sheet = wb['temp']
        except KeyError:
            abort(500, "テンプレートファイルに 'temp' という名前のシートが見つかりません。")

        initial_sheets = wb.sheetnames
        
        # Excelシート名に使用できない文字
        INVALID_TITLE_CHARS = [':', '\\', '/', '?', '*', '[', ']']

        for name in kouban_labels:
            # '000000' は有休用シートとして特別扱いする
            is_yukyu_sheet = (name == '000000')

            # シート名のサニタイズ
            if is_yukyu_sheet:
                safe_name = '有休'
            else:
                safe_name = name
                for char in INVALID_TITLE_CHARS:
                    safe_name = safe_name.replace(char, '_')
                safe_name = safe_name[:31] # 31文字制限

            # シート名が重複しないようにする
            if safe_name in wb.sheetnames:
                safe_name = f"{safe_name[:28]}_{len(wb.sheetnames)}"

            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = safe_name
            
            # --- 5. 各シートにデータを書き込む ---
            category_info = category_info_map.get(name, {})
            
            if is_yukyu_sheet:
                new_sheet['B1'] = '有休'
                new_sheet['B2'] = ''
            else:
                new_sheet['B1'] = name
                new_sheet['B2'] = category_info.get('client', '')
            
            new_sheet.merge_cells('B1:C1')
            # D1 (案件名) は作業者列数が確定してから書き込むため、ここではスキップ
            new_sheet.merge_cells('B2:C2')

            # --- 6. 作業者リストの作成とヘッダーの書き込み ---
            kako_employees = set()
            genba_employees = set()
            employee_data_for_kouban = kouban_data_map.get(name, {})
            for emp_name, date_map in employee_data_for_kouban.items():
                for type_map in date_map.values():
                    if type_map.get('加工', 0) > 0:
                        kako_employees.add(emp_name)
                    if type_map.get('現場', 0) > 0:
                        genba_employees.add(emp_name)
            
            sorted_kako_employees = sorted(list(kako_employees))
            sorted_genba_employees = sorted(list(genba_employees))

            # --- 案件名の書き込みとセル結合 ---
            # 作業者列の総数を計算
            total_worker_cols = len(sorted_kako_employees) + len(sorted_genba_employees)
            
            if is_yukyu_sheet:
                new_sheet['D1'] = ''
            else:
                new_sheet['D1'] = category_info.get('project', '')
            
            # 最低でもD列(4)からM列(13)までは結合する。データがそれ以上に及ぶ場合は、データの最終列まで結合する
            merge_end_col = max(13, 4 + total_worker_cols - 1)
            new_sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=merge_end_col)

            blue_fill = PatternFill(start_color="BDE4F4", end_color="BDE4F4", fill_type="solid")
            green_fill = PatternFill(start_color="D7EAD7", end_color="D7EAD7", fill_type="solid")
            start_col_idx = 4 # D列

            if sorted_kako_employees:
                kako_start_col = start_col_idx
                kako_end_col = kako_start_col + len(sorted_kako_employees) - 1
                # 1セルだけの場合は結合しない（Excelエラー回避）
                if kako_end_col > kako_start_col:
                    new_sheet.merge_cells(start_row=2, start_column=kako_start_col, end_row=2, end_column=kako_end_col)
                header_cell_kako = new_sheet.cell(row=2, column=kako_start_col, value='加工')
                header_cell_kako.fill = blue_fill
                header_cell_kako.alignment = Alignment(horizontal='center')

            if sorted_genba_employees:
                genba_start_col = start_col_idx + len(sorted_kako_employees)
                genba_end_col = genba_start_col + len(sorted_genba_employees) - 1
                # 1セルだけの場合は結合しない（Excelエラー回避）
                if genba_end_col > genba_start_col:
                    new_sheet.merge_cells(start_row=2, start_column=genba_start_col, end_row=2, end_column=genba_end_col)
                header_cell_genba = new_sheet.cell(row=2, column=genba_start_col, value='現場')
                if is_yukyu_sheet:
                    header_cell_genba.value = '時間'
                header_cell_genba.fill = green_fill
                header_cell_genba.alignment = Alignment(horizontal='center')
            
            # 3行目に作業者名を書き込み (D3, E3, ...)
            for i, emp_name in enumerate(sorted_kako_employees):
                # セル内改行のために名前を分割 (全角スペースも考慮)
                name_parts = emp_name.replace('　', ' ').split(' ', 1)
                if len(name_parts) == 2:
                    formatted_name = f"{name_parts[0]}\n{name_parts[1]}"
                else:
                    formatted_name = emp_name
                
                cell = new_sheet.cell(row=3, column=start_col_idx + i, value=formatted_name)
                # セルのスタイルで折り返しを有効にする
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            for i, emp_name in enumerate(sorted_genba_employees):
                name_parts = emp_name.replace('　', ' ').split(' ', 1)
                if len(name_parts) == 2:
                    formatted_name = f"{name_parts[0]}\n{name_parts[1]}"
                else:
                    formatted_name = emp_name
                
                cell = new_sheet.cell(row=3, column=start_col_idx + len(sorted_kako_employees) + i, value=formatted_name)
                cell.alignment = Alignment(wrap_text=True, vertical='top')


            # --- 7. 工数データの書き込み ---
            # テンプレートのA列（5行目〜）から日付（数値）と行番号のマッピングを作成
            date_row_map = {}
            # テンプレートのA列を5行目から40行目まで走査して日付と行のマッピングを作成
            for row in new_sheet.iter_rows(min_row=5, max_row=40, min_col=1, max_col=1):
                cell = row[0]
                try:
                    # セルの値が文字列の数字である可能性も考慮してintに変換
                    day_val_in_cell = int(cell.value)
                    date_row_map[day_val_in_cell] = cell.row
                except (ValueError, TypeError):
                    # intに変換できない値は無視
                    continue

            day_to_datestr_map = {}
            d = start_date
            while d <= end_date:
                day_to_datestr_map[d.day] = d.strftime('%Y-%m-%d')
                d += timedelta(days=1)

            for day_val, target_row in date_row_map.items():
                date_str = day_to_datestr_map.get(day_val)
                if not date_str:
                    continue

                daily_kako_total_minutes = 0
                daily_genba_total_minutes = 0

                for i, emp_name in enumerate(sorted_kako_employees):
                    minutes = employee_data_for_kouban.get(emp_name, {}).get(date_str, {}).get('加工', 0)
                    daily_kako_total_minutes += minutes
                    if minutes > 0:
                        new_sheet.cell(row=target_row, column=start_col_idx + i, value=minutes / 60.0)

                for i, emp_name in enumerate(sorted_genba_employees):
                    minutes = employee_data_for_kouban.get(emp_name, {}).get(date_str, {}).get('現場', 0)
                    daily_genba_total_minutes += minutes
                    if minutes > 0:
                        new_sheet.cell(row=target_row, column=start_col_idx + len(sorted_kako_employees) + i, value=minutes / 60.0)

                if daily_kako_total_minutes > 0:
                    new_sheet.cell(row=target_row, column=2, value=daily_kako_total_minutes / 60.0)

                if daily_genba_total_minutes > 0:
                    new_sheet.cell(row=target_row, column=3, value=daily_genba_total_minutes / 60.0)

            # --- 8. 不要な列の削除 ---
            # 使用した最終列(merge_end_col)より右側の列を削除して、テンプレートに残っている書式等をクリアする
            try:
                delete_start_col = merge_end_col + 1
                current_max_col = new_sheet.max_column
                if current_max_col >= delete_start_col:
                    new_sheet.delete_cols(delete_start_col, current_max_col - delete_start_col + 1)
            except Exception as e:
                current_app.logger.warning(f"Failed to delete columns in sheet '{safe_name}': {e}")

        for sheet_name in initial_sheets:
            try:
                wb.remove(wb[sheet_name])
            except Exception as e:
                current_app.logger.warning(f"Failed to remove template sheet '{sheet_name}': {e}")

        # --- 7. Excelファイルを生成して返す ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.read()
        b64_data = base64.b64encode(excel_data).decode('utf-8')

        file_name = f"{end_date.strftime('%Y年%m月度')}_工番別集計表.xlsx"
        return jsonify({"file_name": file_name, "file_content": b64_data}), 200

    except Exception as e:
        current_app.logger.error(f"Excel generation failed: {e}", exc_info=True)
        abort(500, f"Excel生成中にエラーが発生しました: {str(e)}")

@api_bp.route("/manager/staff-summary/excel", methods=["POST"])
@token_required
@login_required
def download_staff_summary_excel():
    """
    スタッフ別集計表のExcelファイルを生成して返す。
    """
    try:
        data = request.get_json() or {}
        target_month = data.get('target_month', 'current') # 'current' or 'previous'

        # 1. 集計期間の決定
        jst = timezone(timedelta(hours=9))
        base_date = datetime.now(jst)
        start_date, end_date = calculate_monthly_period(base_date)

        if target_month == 'previous':
            prev_month_base = start_date - timedelta(days=1)
            start_date, end_date = calculate_monthly_period(prev_month_base)

        current_app.logger.info(f"Staff summary Excel period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        # 2. 全アクティブ従業員を取得
        mappings_ref = db.collection('employee_mappings')
        query_mappings = mappings_ref.where(filter=FieldFilter("status", "==", "active"))
        mappings_docs = query_mappings.stream()
        
        employees = []
        for doc in mappings_docs:
            d = doc.to_dict()
            employees.append({
                'id': doc.id, # company_employee_id
                'name': d.get('name', 'Unknown')
            })
        employees.sort(key=lambda x: x['id']) # ID順

        # 3. 日報データを取得
        source_collection = _resolve_summary_source_collection(target_month, start_date, end_date, "enj")
        current_app.logger.info(
            f"staff-summary/excel source collection for '{target_month}': {source_collection}"
        )
        query = db.collection(source_collection).where(filter=FieldFilter("date", ">=", start_date)).where(filter=FieldFilter("date", "<=", end_date))
        docs = query.stream()

        # data_map[employee_id][date_str][category_a_id] = minutes
        data_map = {}
        # data_map_b[employee_id][date_str][category_b_label] = minutes
        data_map_b = {}
        # employee_projects[employee_id] = set(category_b_labels)
        employee_projects = {}
        
        for doc in docs:
            rep = doc.to_dict()
            emp_id = rep.get('company_employee_id')
            if not emp_id: continue
            
            # Timestamp to string
            r_date = rep.get('date')
            if isinstance(r_date, datetime):
                date_str = r_date.strftime('%Y-%m-%d')
            else:
                continue

            tasks = rep.get('tasks', [])
            
            if emp_id not in data_map: data_map[emp_id] = {}
            if date_str not in data_map[emp_id]: data_map[emp_id][date_str] = {}
            
            if emp_id not in data_map_b: data_map_b[emp_id] = {}
            if date_str not in data_map_b[emp_id]: data_map_b[emp_id][date_str] = {}
            if emp_id not in employee_projects: employee_projects[emp_id] = set()
            
            for task in tasks:
                cat_a = task.get('categoryA_id')
                cat_b_label = task.get('categoryB_label')
                time_min = task.get('time', 0)
                
                if cat_a:
                    # A99（有休・忌引）を有休列（A00）と同一集計
                    agg_cat_a = 'A00' if cat_a == 'A99' else cat_a
                    data_map[emp_id][date_str][agg_cat_a] = data_map[emp_id][date_str].get(agg_cat_a, 0) + time_min
                
                # 工事番号別集計 (実績がある場合のみ)
                if cat_b_label and time_min > 0:
                    data_map_b[emp_id][date_str][cat_b_label] = data_map_b[emp_id][date_str].get(cat_b_label, 0) + time_min
                    employee_projects[emp_id].add(cat_b_label)

        # 4. Excel生成
        # テンプレートファイル名: template_staff.xlsx (openpyxl対応のため拡張子を変更して想定)
        template_path = os.path.join(current_app.root_path, 'temp', 'template_staff.xlsx')
        if not os.path.exists(template_path):
             abort(500, f"テンプレートファイルが見つかりません: {template_path}")

        wb = openpyxl.load_workbook(template_path)
        try:
            template_sheet = wb['temp']
        except KeyError:
            # シート名が 'temp' でない場合はアクティブシートを使用
            template_sheet = wb.active
        
        # 集計する業務種別IDの固定順序（テンプレ C 列〜の見出しと一致させる。A11/A12 追加に伴い工事番号列は +2）
        cat_a_order = [
            'A01', 'A11', 'A02', 'A05', 'A10', 'A03', 'A12', 'A09', 'A04', 'A07', 'A06', 'A08', 'A00',
        ]
        # Category B（工事番号）動的列の開始列（1-based）。旧15列目→A列2列増に合わせ17列目（Q）以降
        staff_summary_cat_b_start_col = 17
        
        # 各従業員のシートを作成
        for emp in employees:
            emp_id = emp['id']
            emp_name = emp['name']
            
            # シートをコピー
            ws = wb.copy_worksheet(template_sheet)
            
            # シート名 (使用不可文字を除去)
            safe_name = emp_name
            for char in [':', '\\', '/', '?', '*', '[', ']']:
                safe_name = safe_name.replace(char, '')
            ws.title = safe_name[:30] # 31文字制限
            
            # ヘッダー情報
            ws['B1'] = emp_name
            ws['B2'] = emp_id
            
            # Category B (工事番号) ヘッダー (Q列=17列目以降)
            emp_proj_list = sorted(list(employee_projects.get(emp_id, set())))
            for i, proj_label in enumerate(emp_proj_list):
                col_idx = staff_summary_cat_b_start_col + i
                ws.cell(row=2, column=col_idx, value=proj_label).alignment = Alignment(horizontal='center')
            
            # データ行の書き込み (行3〜33)
            curr = start_date
            row_idx = 3
            
            while curr <= end_date and row_idx <= 33:
                d_str = curr.strftime('%Y-%m-%d')
                day_data = data_map.get(emp_id, {}).get(d_str, {})
                day_data_b = data_map_b.get(emp_id, {}).get(d_str, {})
                
                row_total_min = 0
                
                # C列(3) 〜 O列(15) … 業務種別13列
                for i, cat_id in enumerate(cat_a_order):
                    col_idx = 3 + i
                    val_min = day_data.get(cat_id, 0)
                    row_total_min += val_min
                    
                    if val_min > 0:
                        # 時間単位で出力 (例: 90分 -> 1.5)
                        ws.cell(row=row_idx, column=col_idx, value=val_min/60.0)
                
                # B列(2) 合計
                if row_total_min > 0:
                    ws.cell(row=row_idx, column=2, value=row_total_min/60.0)
                
                # Q列(17) 〜 (Category B / 工事番号)
                for i, proj_label in enumerate(emp_proj_list):
                    col_idx = staff_summary_cat_b_start_col + i
                    val_min = day_data_b.get(proj_label, 0)
                    
                    if val_min > 0:
                        ws.cell(row=row_idx, column=col_idx, value=val_min/60.0)
                
                curr += timedelta(days=1)
                row_idx += 1
        
        # テンプレートシートを削除
        if template_sheet in wb._sheets:
            wb.remove(template_sheet)
        
        # 保存してBase64化
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64_data = base64.b64encode(output.read()).decode('utf-8')
        
        file_name = f"スタッフ別集計_{end_date.strftime('%Y年%m月度')}.xlsx"
        return jsonify({"file_name": file_name, "file_content": b64_data}), 200

    except Exception as e:
        current_app.logger.error(f"Staff summary Excel generation failed: {e}", exc_info=True)
        abort(500, f"Excel生成中にエラーが発生しました: {str(e)}")


def _parse_daily_report_doc_id(doc_id: str) -> tuple[str | None, str | None]:
    """{company_employee_id}_{YYYY-MM-DD} 形式から ID と日付を返す。"""
    if not doc_id or "_" not in doc_id:
        return None, None
    emp_id, date_part = doc_id.rsplit("_", 1)
    if len(date_part) != 10 or date_part[4] != "-" or date_part[7] != "-":
        return None, None
    return emp_id, date_part


def _report_date_field_to_str(r_date) -> str | None:
    if r_date is None:
        return None
    if isinstance(r_date, datetime):
        return r_date.strftime("%Y-%m-%d")
    if isinstance(r_date, str) and len(r_date) >= 10:
        return r_date[:10]
    return None


def _iter_tasks_for_net_csv(tasks):
    """tasks が list または map（連番キー）のどちらでも1要素ずつ返す。"""
    if tasks is None:
        return
    if isinstance(tasks, list):
        for t in tasks:
            yield t
    elif isinstance(tasks, dict):
        for t in tasks.values():
            yield t


def _norm_str_id(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _safe_int_minutes(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _net_task_comment_for_csv(task: dict) -> str:
    """ネット日報タスクの comment を CSV 用に正規化（空は空文字）。"""
    if not isinstance(task, dict):
        return ""
    c = task.get("comment")
    if c is None:
        return ""
    return str(c).strip()


def _is_net_daily_report_group(group_id) -> bool:
    """ネット事業部の日報（group_id=3）のみ。"""
    if group_id is None:
        return False
    return str(group_id).strip() == "3"


def _net_staff_summary_label_key(label: str | None) -> str:
    return (label or "").strip()


def _format_net_staff_summary_cat_a_display_label(label: str | None) -> str:
    """業務種別を Excel 表示用に整形する。全角括弧 （…） がある場合はそのブロックを除く。"""
    s = (label or "").strip()
    if not s:
        return ""
    s = re.sub(r"（[^）]*）", "", s)
    return s.strip()


def _normalize_hex_rgb_for_openpyxl_fill(color_raw: str | None) -> str | None:
    """category_a_settings の色文字列を openpyxl PatternFill 用 6 桁 RGB に正規化。"""
    if color_raw is None:
        return None
    s = str(color_raw).strip()
    if not s:
        return None
    if s.startswith("#"):
        s = s[1:]
    if len(s) == 3 and all(c in "0123456789ABCDEFabcdef" for c in s):
        s = "".join(c * 2 for c in s)
    if len(s) == 6 and all(c in "0123456789ABCDEFabcdef" for c in s):
        return s.upper()
    return None


# スタッフ別（ネット）Excel: ヘッダ行のみ 9pt、データ行は 8pt（Meiryo）
_NET_STAFF_SUMMARY_FONT_HEADER = Font(name="Meiryo", size=9)
_NET_STAFF_SUMMARY_FONT_BODY = Font(name="Meiryo", size=8)


def _apply_net_staff_summary_sheet_font(ws, *, last_row: int, last_col: int = 6) -> None:
    """シート内の使用矩形にフォントを適用（1〜2行目 9pt、3行目以降 8pt）。"""
    for r in range(1, last_row + 1):
        font = _NET_STAFF_SUMMARY_FONT_HEADER if r <= 2 else _NET_STAFF_SUMMARY_FONT_BODY
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).font = font


def _lookup_category_a_color_from_settings(category_a_settings: dict | None, category_a_id: str | None) -> str | None:
    """category_b.category_a_settings から業務種別 ID に対応する背景色を取得。"""
    if not category_a_settings or category_a_id is None:
        return None
    cid = str(category_a_id).strip()
    if not cid:
        return None
    raw = category_a_settings.get(cid)
    if raw is None:
        raw = category_a_settings.get(category_a_id)
    if isinstance(raw, dict):
        raw = raw.get("color") or raw.get("bg") or raw.get("hex")
    return _normalize_hex_rgb_for_openpyxl_fill(raw)


def _fetch_net_category_a_for_staff_summary_excel() -> list[dict]:
    """ネット向け category_a（アクティブ）を order・label でソート。"""
    rows: list[dict] = []
    for doc in db.collection("category_a").where(filter=FieldFilter("kind", "==", "net")).stream():
        data = doc.to_dict()
        if data.get("active") is False:
            continue
        rows.append(
            {
                "id": doc.id,
                "label": data.get("label", ""),
                "order": data.get("order", 9999),
            }
        )
    rows.sort(key=lambda x: (x["order"], _net_staff_summary_label_key(x["label"])))
    return [{"id": x["id"], "label": x["label"]} for x in rows]


def _fetch_net_category_b_for_staff_summary_excel() -> list[dict]:
    """ネット向け category_b（アクティブ）を order 昇順でソート。"""
    cats: list[dict] = []
    for doc in db.collection("category_b").where(filter=FieldFilter("kind", "==", "net")).stream():
        data = doc.to_dict()
        if data.get("active") is False:
            continue
        cats.append(
            {
                "id": doc.id,
                "label": data.get("label", ""),
                "order": data.get("order", 9999),
                "category_a_settings": data.get("category_a_settings") or {},
            }
        )
    cats.sort(key=lambda x: (x["order"], _net_staff_summary_label_key(x["label"])))
    return [
        {"id": x["id"], "label": x["label"], "category_a_settings": x["category_a_settings"]}
        for x in cats
    ]


def _parse_user_entered_day_date(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None
    return None


def _years_since_entered_day(entered_day: datetime | None, as_of: datetime) -> int:
    if entered_day is None:
        return 0
    years = as_of.year - entered_day.year
    if (as_of.month, as_of.day) < (entered_day.month, entered_day.day):
        years -= 1
    return max(years, 0)


def _net_work_kind_labor_factor(work_kind_raw) -> Decimal:
    wk = _norm_str_id(work_kind_raw)
    if wk == "3":
        return Decimal("1.0")
    if wk == "10":
        return Decimal("0.75")
    return Decimal("0.67")


def _round_to_nearest_10_half_up(v: Decimal) -> int:
    return int((v / Decimal("10")).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * Decimal("10"))


def _get_net_summary_officer_overrides() -> tuple[str | None, str | None, str | None]:
    """
    ネット集計でのみ使う執行役員の上書き設定。
    import 失敗時は API 全体を巻き込まないよう None を返す。
    """
    try:
        from app_core import config as _cfg
        officer_id = getattr(_cfg, "OFFICER_ID", None)
        officer_entered_day = getattr(_cfg, "OFFICER_ENTERED_DAY", None)
        officer_name = getattr(_cfg, "OFFICER_NAME", None)
        return (
            _norm_str_id(officer_id),
            _norm_str_id(officer_entered_day),
            _norm_str_id(officer_name),
        )
    except Exception:
        return None, None, None


def _calc_net_staff_monthly_labor_cost(user_data: dict, as_of: datetime, employee_id: str | None = None) -> int:
    entered_day = _parse_user_entered_day_date(user_data.get("entered_day"))
    work_kind_raw = user_data.get("work_kind_id") or user_data.get("work_kind")
    officer_id, officer_entered_day_raw, _ = _get_net_summary_officer_overrides()
    # OFFICER_ID は int / str / "210501.0" など表現ゆれがあり得るため、
    # Jobcan 突合せと同じ正規化で比較する。
    normalized_emp = _normalize_jobcan_employee_id_for_match(employee_id)
    normalized_officer = _normalize_jobcan_employee_id_for_match(officer_id)
    if normalized_emp and normalized_emp == normalized_officer:
        officer_entered_day = _parse_user_entered_day_date(officer_entered_day_raw)
        if officer_entered_day is not None:
            entered_day = officer_entered_day
        # 執行役員は work_kind を固定で 3 扱いにする
        work_kind_raw = "3"
    years = _years_since_entered_day(entered_day, as_of)
    # CAREER_COEFFICIENT は 0.03 形式を想定。
    # もし 1.03 / 1.025 のような値が入っていても、加算率としては 0.03 / 0.025 として扱う。
    career_increment = Decimal(str(CAREER_COEFFICIENT))
    if career_increment >= Decimal("1"):
        career_increment = career_increment - Decimal("1")
    tenure_factor = Decimal("1") + (career_increment * Decimal(years))
    work_kind_factor = _net_work_kind_labor_factor(work_kind_raw)
    cost = Decimal(str(MONTHLY_lABOR_COSTS_NET)) * tenure_factor * work_kind_factor
    return _round_to_nearest_10_half_up(cost)


def _fetch_net_staff_for_summary_blocks(start_date: datetime, end_date: datetime, as_of: datetime) -> list[dict]:
    """
    対象月度の daily_reports（ネット日報）に出現する社員を抽出し、
    工務スタッフ別と同様に employee_mappings の name を表示名として返す。
    """
    query = (
        db.collection(COLLECTION_DAILY_REPORTS)
        .where(filter=FieldFilter("date", ">=", start_date))
        .where(filter=FieldFilter("date", "<=", end_date))
    )

    target_emp_ids: set[str] = set()
    for doc in query.stream():
        rep = doc.to_dict() or {}
        if not _is_net_daily_report_group(rep.get("group_id")):
            continue
        parsed_e, _ = _parse_daily_report_doc_id(doc.id)
        emp_id = _norm_str_id(rep.get("company_employee_id")) or parsed_e
        if emp_id:
            target_emp_ids.add(emp_id)

    if not target_emp_ids:
        return []

    mapping_docs = (
        db.collection("employee_mappings")
        .where(filter=FieldFilter("status", "==", "active"))
        .stream()
    )
    name_map: dict[str, str] = {}
    for doc in mapping_docs:
        if doc.id in target_emp_ids:
            d = doc.to_dict() or {}
            name_map[doc.id] = d.get("name", "Unknown")

    staff = []
    for emp_id in sorted(target_emp_ids):
        user_snap = db.collection("users").document(emp_id).get()
        user_data = user_snap.to_dict() if user_snap.exists else {}
        staff.append(
            {
                "id": emp_id,
                "name": name_map.get(emp_id, "Unknown"),
                "labor_cost": _calc_net_staff_monthly_labor_cost(user_data or {}, as_of, emp_id),
            }
        )
    return staff


def _aggregate_net_staff_task_minutes_for_summary(start_date: datetime, end_date: datetime) -> dict[str, dict[tuple[str, str], int]]:
    """
    対象月度のネット日報から、社員×(category_b_id, category_a_id) ごとの分数を合算して返す。
    """
    query = (
        db.collection(COLLECTION_DAILY_REPORTS)
        .where(filter=FieldFilter("date", ">=", start_date))
        .where(filter=FieldFilter("date", "<=", end_date))
    )

    agg: dict[str, dict[tuple[str, str], int]] = {}
    for doc in query.stream():
        rep = doc.to_dict() or {}
        if not _is_net_daily_report_group(rep.get("group_id")):
            continue

        parsed_e, _ = _parse_daily_report_doc_id(doc.id)
        emp_id = _norm_str_id(rep.get("company_employee_id")) or parsed_e
        if not emp_id:
            continue

        tasks = rep.get("tasks")
        for task in _iter_tasks_for_net_csv(tasks):
            if not isinstance(task, dict):
                continue
            cat_a_id = _norm_str_id(task.get("categoryA_id"))
            cat_b_id = _norm_str_id(task.get("categoryB_id"))
            mins = _safe_int_minutes(task.get("time")) or 0
            if not cat_a_id or not cat_b_id or mins <= 0:
                continue

            by_key = agg.setdefault(emp_id, {})
            key = (cat_b_id, cat_a_id)
            by_key[key] = by_key.get(key, 0) + mins
    return agg


def _build_net_staff_summary_excel_workbook(start_date: datetime, end_date: datetime, as_of: datetime) -> openpyxl.Workbook:
    """
    スタッフ別（ネット）Excelのレイアウト骨子。
    - シート名: YYYY年MM月度（月度終了日ベース）
    - 列 D〜F: カテゴリは列挙しない（プレースホルダのみ。例: E1=合計、D2=累計人件費）
    - 列 A〜C: 行3〜 に左エリア。列Cに業務種別を縦に（集計項目ブロック数ぶん繰り返し）。
      集計項目の並びは category_b.order 昇順。
      列Bは業務種別が「全般」の行のみ集計項目ラベル、それ以外は空。
      列Aはその行の集計項目ブロックに対応する category_b.category_a_settings のカラーマップで、
      該当業務種別 ID のセルを塗りつぶす（値は入れない）。
    """

    cat_a_list = _fetch_net_category_a_for_staff_summary_excel()
    cat_b_list = _fetch_net_category_b_for_staff_summary_excel()
    staff_list = _fetch_net_staff_for_summary_blocks(start_date, end_date, as_of)
    task_minutes_map = _aggregate_net_staff_task_minutes_for_summary(start_date, end_date)
    staff_total_task_minutes = {
        s["id"]: sum(task_minutes_map.get(s["id"], {}).values())
        for s in staff_list
    }
    all_staff_total_task_minutes = sum(staff_total_task_minutes.values())
    if not cat_a_list:
        abort(400, "ネット向け業務種別（category_a）が取得できません。")
    if not cat_b_list:
        abort(400, "ネット向け集計項目（category_b）が取得できません。")

    na = len(cat_a_list)
    nb = len(cat_b_list)
    zenpan_key = _net_staff_summary_label_key("全般")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = end_date.strftime("%Y年%m月度")

    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- 右側 D〜F: カテゴリ列挙なし（プレースホルダ）---
    ws.cell(row=1, column=5).value = "合計"
    ws.cell(row=1, column=5).alignment = hdr_align
    total_labor_cost = sum(int(s.get("labor_cost") or 0) for s in staff_list)
    total_cost_cell = ws.cell(row=2, column=4)
    total_cost_cell.value = total_labor_cost
    total_cost_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    total_cost_cell.number_format = "#,##0"

    # --- G列以降: スタッフ別 3列1ブロック（中央列1行目=スタッフ名、左列2行目=経費額） ---
    staff_block_start_col = 7  # G
    for i, staff in enumerate(staff_list):
        block_left = staff_block_start_col + i * 3
        block_middle = block_left + 1
        staff_name = staff["name"]
        officer_id, _, officer_name = _get_net_summary_officer_overrides()
        if (
            officer_name
            and _normalize_jobcan_employee_id_for_match(staff.get("id")) == _normalize_jobcan_employee_id_for_match(officer_id)
        ):
            staff_name = officer_name
        ws.cell(row=1, column=block_middle).value = staff_name
        ws.cell(row=1, column=block_middle).alignment = hdr_align

        cost_cell = ws.cell(row=2, column=block_left)
        cost_cell.value = staff["labor_cost"]
        cost_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        cost_cell.number_format = "#,##0"

    # --- 左3列: 行3〜 縦リスト（列C=業務種別繰り返し、列B=全般行のみ集計項目名）---
    total_left_rows = na * nb
    for i in range(total_left_rows):
        r = 3 + i
        bi = i // na
        ai = i % na
        a_label = cat_a_list[ai]["label"]
        a_display = _format_net_staff_summary_cat_a_display_label(a_label)
        ws.cell(row=r, column=3).value = a_display
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if _net_staff_summary_label_key(a_display) == zenpan_key:
            ws.cell(row=r, column=2).value = cat_b_list[bi]["label"]
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        a_cell = ws.cell(row=r, column=1)
        a_cell.value = None
        settings = cat_b_list[bi].get("category_a_settings") or {}
        a_id = cat_a_list[ai]["id"]
        hex_rgb = _lookup_category_a_color_from_settings(settings, a_id)
        if hex_rgb:
            a_cell.fill = PatternFill(fill_type="solid", fgColor=hex_rgb)

        # スタッフ別ブロック:
        # - 中央列: 当該行（category_b × category_a）の月度合計時間（小数第2位）
        # - 右列: そのスタッフの月度トータル勤務（タスク）時間に対する割合
        # - 左列: スタッフ別人件費 × 割合（当該業務にかかった経費）
        b_id = cat_b_list[bi]["id"]
        per_staff_minutes: list[int] = []
        for staff in staff_list:
            mins = task_minutes_map.get(staff["id"], {}).get((b_id, a_id), 0)
            per_staff_minutes.append(mins)
        row_total_minutes_all_staff = sum(per_staff_minutes)

        # 合計ブロック（D:E:F）
        # - E列: タスク別累計時間（全スタッフ合算）
        # - F列: タスク別累計時間 / 全スタッフ月度トータル勤務時間
        # - D列: 全スタッフ累計人件費 × F列（タスク別経費）
        if row_total_minutes_all_staff > 0:
            total_hours = (Decimal(row_total_minutes_all_staff) / Decimal("60")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            total_hours_cell = ws.cell(row=r, column=5)
            total_hours_cell.value = float(total_hours)
            total_hours_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            total_hours_cell.number_format = "0.00"

            if all_staff_total_task_minutes > 0:
                total_ratio = Decimal(row_total_minutes_all_staff) / Decimal(all_staff_total_task_minutes)
                total_ratio_cell = ws.cell(row=r, column=6)
                total_ratio_cell.value = float(total_ratio)
                total_ratio_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                total_ratio_cell.number_format = "0.00%"

                total_alloc_cost = Decimal(total_labor_cost) * total_ratio
                total_alloc_cost_cell = ws.cell(row=r, column=4)
                total_alloc_cost_cell.value = int(total_alloc_cost.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                total_alloc_cost_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                total_alloc_cost_cell.number_format = "#,##0"

        for si, staff in enumerate(staff_list):
            mins = per_staff_minutes[si]
            block_left = 7 + si * 3
            block_middle = block_left + 1
            block_right = block_left + 2

            if mins > 0:
                hours = (Decimal(mins) / Decimal("60")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                hours_cell = ws.cell(row=r, column=block_middle)
                hours_cell.value = float(hours)
                hours_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                hours_cell.number_format = "0.00"

            staff_total_minutes = int(staff_total_task_minutes.get(staff["id"], 0) or 0)
            if staff_total_minutes > 0 and mins > 0:
                ratio = Decimal(mins) / Decimal(staff_total_minutes)
                ratio_cell = ws.cell(row=r, column=block_right)
                ratio_cell.value = float(ratio)
                ratio_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                ratio_cell.number_format = "0.00%"

                alloc_cost = Decimal(int(staff.get("labor_cost") or 0)) * ratio
                alloc_cost_cell = ws.cell(row=r, column=block_left)
                alloc_cost_cell.value = int(alloc_cost.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                alloc_cost_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                alloc_cost_cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    for i in range(len(staff_list)):
        block_left = 7 + i * 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(block_left)].width = 9
        ws.column_dimensions[openpyxl.utils.get_column_letter(block_left + 1)].width = 9
        ws.column_dimensions[openpyxl.utils.get_column_letter(block_left + 2)].width = 9

    last_data_row = max(2 + na * nb, 2)
    total_row = last_data_row + 1
    last_col = max(6, 6 + len(staff_list) * 3)

    # D列以降のみ、3行目〜末尾データ行の合計を最終行へ追加
    if last_data_row >= 3:
        for c in range(4, last_col + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            ws.cell(row=total_row, column=c).value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"

    for r in range(1, total_row + 1):
        ws.row_dimensions[r].height = 15

    # ウィンドウ枠の固定: 1〜2 行目・A〜F 列まで（スクロール領域は G3 左上）
    ws.freeze_panes = "G3"

    _apply_net_staff_summary_sheet_font(ws, last_row=total_row, last_col=last_col)

    return wb


@api_bp.route("/manager/net-staff-summary/excel", methods=["POST"])
@token_required
@login_required
def download_net_staff_summary_excel_placeholder():
    """
    ネット事業部「スタッフ別」集計 Excel。
    テンプレートは使わずプログラム生成。ヘッダレイアウトのみ先行実装し、数値集計は後続。
    """
    try:
        data = request.get_json() or {}
        target_month = data.get("target_month", "current")

        jst = timezone(timedelta(hours=9))
        base_date = datetime.now(jst)
        start_date, end_date = calculate_monthly_period(base_date)

        if target_month == "previous":
            prev_month_base = start_date - timedelta(days=1)
            start_date, end_date = calculate_monthly_period(prev_month_base)

        current_app.logger.info(
            "net-staff-summary/excel period: %s to %s",
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d"),
        )

        wb = _build_net_staff_summary_excel_workbook(start_date, end_date, base_date)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.read()
        b64_data = base64.b64encode(excel_data).decode("utf-8")

        file_name = f"{end_date.strftime('%Y年%m月度')}_スタッフ別（ネット）.xlsx"
        return jsonify({"file_name": file_name, "file_content": b64_data}), 200
    except Exception as e:
        current_app.logger.error(f"net-staff-summary/excel failed: {e}", exc_info=True)
        abort(500, f"Excel生成中にエラーが発生しました: {str(e)}")


@api_bp.route("/manager/net-task-summary/csv", methods=["POST"])
@token_required
@login_required
def download_net_task_summary_csv():
    """
    ネット事業部の日報データを月度単位で集計し、ピボット用の縦持ちCSVを返す。
    集計キー: (company_employee_id, date, categoryA_id, categoryB_id) ／ time は分で合算。
    comment は同一キーに複数タスクがある場合、重複を除き「 | 」で連結する。
    """
    try:
        data = request.get_json() or {}
        target_month = data.get("target_month", "current")

        jst = timezone(timedelta(hours=9))
        base_date = datetime.now(jst)
        start_date, end_date = calculate_monthly_period(base_date)

        if target_month == "previous":
            prev_month_base = start_date - timedelta(days=1)
            start_date, end_date = calculate_monthly_period(prev_month_base)

        current_app.logger.info(
            f"Net task summary CSV period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        )

        source_collection = _resolve_summary_source_collection(target_month, start_date, end_date, "net")
        current_app.logger.info(
            f"net-task-summary/csv source collection for '{target_month}': {source_collection}"
        )

        query = (
            db.collection(source_collection)
            .where(filter=FieldFilter("date", ">=", start_date))
            .where(filter=FieldFilter("date", "<=", end_date))
        )
        docs = query.stream()

        # key: (emp_id, date_str, cat_a_id, cat_b_id)
        # value: { time, categoryA_label, categoryB_label, employee_name, comments(list[str]) }
        agg: dict = {}

        for doc in docs:
            rep = doc.to_dict()
            if not _is_net_daily_report_group(rep.get("group_id")):
                continue

            parsed_e, parsed_d = _parse_daily_report_doc_id(doc.id)
            emp_id = _norm_str_id(rep.get("company_employee_id")) or parsed_e
            if not emp_id:
                continue

            date_str = _report_date_field_to_str(rep.get("date")) or parsed_d
            if not date_str:
                continue

            employee_name = rep.get("employee_name") or ""

            for task in _iter_tasks_for_net_csv(rep.get("tasks")):
                if not isinstance(task, dict):
                    continue
                cat_a = _norm_str_id(task.get("categoryA_id"))
                cat_b = _norm_str_id(task.get("categoryB_id"))
                tmin = _safe_int_minutes(task.get("time"))
                if tmin is None:
                    continue
                if not cat_a or not cat_b:
                    continue

                cmt = _net_task_comment_for_csv(task)
                key = (emp_id, date_str, cat_a, cat_b)
                if key not in agg:
                    agg[key] = {
                        "time": tmin,
                        "categoryA_label": str(task.get("categoryA_label") or ""),
                        "categoryB_label": str(task.get("categoryB_label") or ""),
                        "employee_name": employee_name,
                        "comments": [cmt] if cmt else [],
                    }
                else:
                    agg[key]["time"] += tmin
                    if cmt and cmt not in agg[key]["comments"]:
                        agg[key]["comments"].append(cmt)

        # 出力行のソート（安定したピボット向け）
        rows = []
        for (emp_id, date_str, cat_a, cat_b), v in sorted(
            agg.items(), key=lambda x: (x[0][0], x[0][1], x[0][2], x[0][3])
        ):
            comment_cell = " | ".join(v["comments"]) if v.get("comments") else ""
            rows.append(
                [
                    emp_id,
                    v["employee_name"],
                    date_str,
                    cat_a,
                    v["categoryA_label"],
                    cat_b,
                    v["categoryB_label"],
                    int(v["time"]),
                    comment_cell,
                ]
            )

        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerow(
            [
                "company_employee_id",
                "employee_name",
                "date",
                "categoryA_id",
                "categoryA_label",
                "categoryB_id",
                "categoryB_label",
                "time",
                "comment",
            ]
        )
        writer.writerows(rows)
        csv_bytes = buf.getvalue().encode("utf-8-sig")
        b64_data = base64.b64encode(csv_bytes).decode("utf-8")

        file_name = f"業務別集計_ネット_{end_date.strftime('%Y年%m月度')}.csv"
        return jsonify({"file_name": file_name, "file_content": b64_data}), 200

    except Exception as e:
        current_app.logger.error(f"Net task summary CSV failed: {e}", exc_info=True)
        abort(500, f"CSV生成中にエラーが発生しました: {str(e)}")


def _date_to_allowance_row(day: int) -> int:
    """対象期間の日付の「日」から、テンプレートの行番号(1-based)を返す。21〜31→2〜12、1〜20→13〜32。"""
    if 21 <= day <= 31:
        return 2 + (day - 21)
    if 1 <= day <= 20:
        return 13 + (day - 1)
    return 0


def _calculate_allowance_period_date_only(target_month: str, now_jst: datetime) -> tuple[datetime, datetime]:
    """
    宿泊/現場(全社)Excel専用の期間計算。
    月度の開始・終了日は既存 calculate_monthly_period を流用しつつ、時刻は日付境界に正規化する。
    - start_date: 00:00:00.000000
    - end_date:   23:59:59.999999

    NOTE:
    calculate_monthly_period は「日」のみを切り替える関数で、時刻は保持される。
    そのため datetime.now(...) をそのまま渡すと、例えば start_date が 2026-02-21 14:37 になり、
    2026-02-21 09:00 の日報が「月度先頭日なのに範囲外」になるケースがある。
    月度を日付単位で扱いたい集計では、必ず時刻を日付境界に正規化すること。
    """
    start_date, end_date = calculate_monthly_period(now_jst)
    if target_month == "previous":
        prev_base = start_date - timedelta(days=1)
        start_date, end_date = calculate_monthly_period(prev_base)

    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    return start_date, end_date


@api_bp.route("/manager/allowance/excel", methods=["POST"])
@token_required
@login_required
def download_allowance_excel():
    """
    宿泊/現場(全社)の手当集計Excelを生成する。
    テンプレート: /temp/template_allowance.xlsx
    対象期間: 21日開始〜翌月20日締め。宿泊は get_accommodation_notes_for_employees、現場は get_on_site_status_for_employees を使用。
    """
    try:
        from openpyxl.utils import get_column_letter

        data = request.get_json() or {}
        target_month = data.get("target_month", "current")

        jst = timezone(timedelta(hours=9))
        base_date = datetime.now(jst)
        # NOTE:
        # この集計は「月度を日付単位」で扱う必要があるため、時刻つき now を直接
        # calculate_monthly_period に渡さない。専用関数で 00:00〜23:59:59.999999 に正規化する。
        # （月度先頭日の朝データが漏れる再発防止）
        start_date, end_date = _calculate_allowance_period_date_only(target_month, base_date)

        # 出力ファイル名: 手当集計_2026年03月度.xlsx
        file_name = f"手当集計_{end_date.strftime('%Y')}年{end_date.strftime('%m')}月度.xlsx"

        # アクティブな従業員を社員ID順で取得
        mappings_ref = db.collection("employee_mappings")
        query_mappings = mappings_ref.where(filter=FieldFilter("status", "==", "active"))
        mappings_docs = list(query_mappings.stream())
        all_employees = [(doc.id, doc.to_dict().get("name", "Unknown")) for doc in mappings_docs]
        all_employees.sort(key=lambda x: x[0])
        employee_ids = [e[0] for e in all_employees]
        id_to_name = {e[0]: e[1] for e in all_employees}

        source_collection = _resolve_summary_source_collection(target_month, start_date, end_date, "all")
        current_app.logger.info(
            f"allowance/excel source collection for '{target_month}': {source_collection}"
        )

        # 宿泊・現場データ取得
        accommodation = get_accommodation_notes_for_employees(employee_ids, start_date, end_date)
        on_site = get_on_site_status_for_employees(
            employee_ids,
            start_date,
            end_date,
            collection_name=source_collection,
        )

        # 宿泊が1件以上ある社員のみ
        accommodation_employees = [eid for eid in employee_ids if accommodation.get(eid)]
        # 現場が1件以上ある社員のみ（0.5 or 1.0 の日が1日以上）
        on_site_employees = [eid for eid in employee_ids if any((on_site.get(eid, {}).get(d) or 0) > 0 for d in (on_site.get(eid, {}) or {}))]

        template_dir = os.path.join(current_app.root_path, "temp")
        template_path = os.path.join(template_dir, "template_allowance.xlsx")
        if not os.path.exists(template_path):
            abort(500, f"テンプレートが見つかりません: {template_path}")

        wb = openpyxl.load_workbook(template_path)
        ws_shukuhaku = wb["宿泊"]
        ws_genba = wb["現場"]

        # 1行目: B1,C1,D1…にスタッフ名。2行目〜32行目: データ（A2=21日…A12=31日、A13=1日…A32=20日）
        def write_sheet(ws, staff_list_emp_ids, value_by_emp_date, value_for_date):
            """staff_list_emp_ids: 列に出す社員IDリスト, value_by_emp_date: {emp_id: {date_str: value}}, value_for_date: 日付→表示値の関数 or 定数"""
            for col_idx, emp_id in enumerate(staff_list_emp_ids, start=2):
                name = id_to_name.get(emp_id, "Unknown")
                ws.cell(row=1, column=col_idx, value=name)
                dates_for_emp = value_by_emp_date.get(emp_id, {})
                for date_str, val in dates_for_emp.items():
                    try:
                        d = datetime.strptime(date_str, "%Y-%m-%d")
                        day = d.day
                        row = _date_to_allowance_row(day)
                        if row == 0:
                            continue
                        # 月末日数を超える日は書かない（29,30,31）
                        if day >= 29:
                            month_len = (d.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                            if day > month_len.day:
                                continue
                        v = value_for_date(val) if callable(value_for_date) else value_for_date
                        if v is None or (isinstance(v, (int, float)) and v == 0):
                            continue
                        ws.cell(row=row, column=col_idx, value=v)
                    except (ValueError, TypeError):
                        continue

        # 宿泊: 該当日は 1.0
        write_sheet(ws_shukuhaku, accommodation_employees, accommodation, 1.0)

        # 現場: 該当日は 0.5 or 1.0（get_on_site_status_for_employees の値そのまま）
        write_sheet(ws_genba, on_site_employees, on_site, lambda v: v)

        # 入力のなかった列を削除（テンプレートは列を多めに用意している前提）
        for sheet, staff_list_ids in [(ws_shukuhaku, accommodation_employees), (ws_genba, on_site_employees)]:
            last_used_col = 1 + len(staff_list_ids)  # A=1, B〜最終スタッフ列
            max_col = sheet.max_column
            if max_col > last_used_col:
                sheet.delete_cols(last_used_col + 1, max_col - last_used_col)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64_data = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"file_name": file_name, "file_content": b64_data}), 200

    except Exception as e:
        current_app.logger.error(f"Allowance Excel generation failed: {e}", exc_info=True)
        abort(500, f"宿泊/現場Excelの生成に失敗しました: {str(e)}")


@api_bp.route("/manager/prepare-report-csv", methods=["POST"])
@token_required
@login_required
def prepare_report_csv():
    """
    管理者向けに、日次レポートのCSVを生成し、一時的にFirestoreに保存してダウンロードURLを返す。
    Request Body: { "date": "YYYY-MM-DD" } (省略時は前日)
    """
    try:
        # 1. リクエストパラメータの取得
        data = request.get_json() or {}
        target_date_str = data.get('date')
        target_group_id = data.get('group_id')
        is_monthly = data.get('is_monthly', False)
        
        jst = timezone(timedelta(hours=9))
        if not target_date_str:
            target_date_str = datetime.now(jst).strftime('%Y-%m-%d')
            
        try:
            target_date_obj = datetime.strptime(target_date_str, '%Y-%m-%d')
        except ValueError:
            abort(400, "Invalid date format. Use YYYY-MM-DD.")

        # 2. Firestoreからレポートを取得
        query = db.collection(COLLECTION_DAILY_REPORTS)
        file_name = ""

        if is_monthly:
            # 月度内一括の場合、対象日付が属する月度の範囲を計算
            start_date, end_date = calculate_monthly_period(target_date_obj)
            
            # 範囲検索
            query = query.where(filter=FieldFilter("date", ">=", start_date)).where(filter=FieldFilter("date", "<=", end_date))
            file_name = f"monthly_report_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.csv"
        else:
            # 通常の日次集計
            query = query.where(filter=FieldFilter("date", "==", target_date_obj))
            file_name = f"daily_report_{target_date_str}.csv"
        
        # グループIDが指定されている場合はフィルタリング
        group_id_list = []
        if target_group_id:
            try:
                # カンマ区切り文字列の場合はリストに変換してIN検索 (例: "4,5,6,7,8")
                if isinstance(target_group_id, str) and ',' in target_group_id:
                    group_id_list = [int(x.strip()) for x in target_group_id.split(',')]
                else:
                    # JobcanのグループIDは数値で保存されているため、intに変換して検索
                    group_id_list = [int(target_group_id)]
            except (ValueError, TypeError):
                abort(400, "Invalid group_id format. Must be an integer or comma-separated integers.")

            # 日次集計(等価検索)の場合は、Firestore側でフィルタリングしてもインデックス不要なため適用する。
            # 月度集計(範囲検索)の場合は、複合インデックス未作成エラーを回避するため、Firestoreクエリには含めない。
            if not is_monthly:
                if len(group_id_list) > 1:
                    query = query.where(filter=FieldFilter("group_id", "in", group_id_list))
                else:
                    query = query.where(filter=FieldFilter("group_id", "==", group_id_list[0]))

        docs = query.stream()
        
        reports = []
        for doc in docs:
            data = doc.to_dict()
            # 月度集計の場合はメモリ内でフィルタリングを行う
            if is_monthly and group_id_list:
                # データにgroup_idがない、またはリストに含まれていない場合はスキップ
                if data.get("group_id") not in group_id_list:
                    continue
            reports.append(data)
            
        # データが0件の場合はCSV生成をスキップしてレスポンスを返す
        if not reports:
            return jsonify({
                "status": "success",
                "count": 0,
                "message": "No data found matching the criteria."
            }), 200

        # 3. CSV生成
        output = io.StringIO()
        # Excelで文字化けしないようにBOMを付与
        output.write('\ufeff')
        writer = csv.writer(output)
        
        # ヘッダー
        writer.writerow(['日付', '社員名', '社員ID', '部署', '勤務時間(分)', '工数合計(分)', '業務内容'])
        
        for report in reports:
            # レポートの日付を使用（Timestamp型の場合は変換）
            report_date = report.get('date')
            if isinstance(report_date, datetime):
                row_date_str = report_date.strftime('%Y-%m-%d')
            else:
                row_date_str = target_date_str

            # 業務内容の整形
            tasks = report.get('tasks') or []
            task_details = [f"[{t.get('categoryA_label', '')}/{t.get('categoryB_label', '')}] {t.get('time',0)}分: {t.get('detail', '')}" for t in tasks]
            tasks_str = "\n".join(task_details)
            
            writer.writerow([
                row_date_str,
                report.get('employee_name', ''),
                report.get('company_employee_id', ''),
                report.get('group_name', ''),
                report.get('jobcan_work_minutes', 0),
                report.get('task_total_minutes', 0),
                tasks_str
            ])
            
        csv_content = output.getvalue()
        output.close()

        # 4. Firestoreに一時保存 (download_links)
        # main.py の download_csv が期待するID形式: {manager_id} (固定)
        manager_id = g.user_info.get('company_employee_id')
        # 日付に依存せず、管理者IDをキーにして常に上書き保存する
        doc_id = str(manager_id)

        db.collection("download_links").document(doc_id).set({
            "manager_id": manager_id,
            "csv_content": csv_content,
            "file_name": file_name,
            "created_at": datetime.now(timezone.utc),
            "target_date": target_date_str
        })

        # 5. ダウンロードURLの生成
        base_url = request.host_url.rstrip('/')
        download_url = f"{base_url}/liff/download/{manager_id}"

        return jsonify({
            "status": "success",
            "download_url": download_url,
            "file_name": file_name,
            "count": len(reports)
        }), 200

    except Exception as e:
        current_app.logger.error(f"Failed to prepare report CSV: {e}")
        abort(500, f"CSV generation failed: {e}")

@api_bp.route("/jobcan/paid-holidays", methods=["GET"])
@token_required
def get_jobcan_paid_holidays():
    """
    Jobcanから休暇「使用」情報を取得するAPI
    """
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        vacation_type = request.args.get('vacation_type', 'paid')

        if not from_date or not to_date:
            abort(400, '期間(from, to)を指定してください。')

        # ログインユーザーのJobcan IDを取得
        user_info = get_user_info_by_line_id(g.line_user_id)
        jobcan_employee_id = user_info.get("jobcan_employee_id")

        current_app.logger.info(f"Fetching paid holidays for employee_id: {jobcan_employee_id} (Period: {from_date} to {to_date})")

        if not jobcan_employee_id:
            # Jobcan IDがないユーザーは有休情報を取得できないため、空のデータを返す
            return jsonify({"use_days": []})

        # JobcanServiceを利用
        from services.jobcan_service import JobcanService
        app_env = os.environ.get("APP_ENV", "development")
        is_sandbox = app_env != "production"

        jobcan_service = JobcanService(
            db=db,
            sandbox=is_sandbox,
            raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES
        )

        # JobcanService経由でデータを取得
        result = jobcan_service.get_paid_holidays(jobcan_employee_id, from_date, to_date, vacation_type)

        if result is None:
             return jsonify({'message': 'Jobcan APIからのデータ取得に失敗しました。'}), 500

        return jsonify(result)

    except Exception as e:
        current_app.logger.error(f"Error in get_jobcan_paid_holidays: {e}")
        abort(500, f'サーバーエラー: {str(e)}')


def _is_effectively_no_jobcan_id(value):
    """
    「Jobcan ID が実質ない」とみなす判定。
    None または空文字（空白のみ含む）のみを「IDなし」とする。
    0 や '0' は有効な従業員IDの可能性があるため「IDあり」と扱い、
    以前まで勤務時間が取得できていたユーザーを誤って全日付ゼロにしない。
    """
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    # 数値 0 などは有効IDの可能性があるため False
    return False


# --- TEMP: Cloud Run ログで Jobcan 有休API・use_id→minutes を確認する用（不要になったら削除） ---
_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG = "[TEMP paid-holiday-sync]"


def _temp_log_jobcan_paid_use_days_response(
    logger,
    holidays_data,
    target_company_id: str,
    target_jobcan_id,
    from_str: str,
    to_str: str,
) -> None:
    """
    GET /holiday/v1/employees/{id}/use-days の戻り値が想定構造かをログに残す。
    """
    if holidays_data is None:
        logger.info(
            f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} Jobcan use-days: response is None "
            f"(company_emp={target_company_id} jobcan_emp={target_jobcan_id} range={from_str}..{to_str})"
        )
        return
    if not isinstance(holidays_data, dict):
        logger.info(
            f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} Jobcan use-days: unexpected type={type(holidays_data).__name__} "
            f"company_emp={target_company_id} jobcan_emp={target_jobcan_id}"
        )
        return

    top_keys = sorted(holidays_data.keys())
    use_days = holidays_data.get("use_days")
    n_blocks = len(use_days) if isinstance(use_days, list) else 0

    paid_snapshots = []
    if isinstance(use_days, list):
        for block in use_days:
            if not isinstance(block, dict):
                continue
            for log in block.get("use_logs") or []:
                if not isinstance(log, dict):
                    continue
                if log.get("detail", {}).get("type") != "paid":
                    continue
                ud = log.get("use_days") or {}
                days_val = 0.0
                if isinstance(ud, dict):
                    try:
                        days_val = float(ud.get("days", 0) or 0)
                    except (TypeError, ValueError):
                        days_val = 0.0
                paid_snapshots.append(
                    {
                        "use_date": log.get("use_date"),
                        "use_id": log.get("use_id"),
                        "use_days_days": days_val,
                        "detail_type": (log.get("detail") or {}).get("type"),
                    }
                )

    try:
        payload = json.dumps(paid_snapshots, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        payload = str(paid_snapshots)

    logger.info(
        f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} Jobcan use-days: company_emp={target_company_id} "
        f"jobcan_emp={target_jobcan_id} range={from_str}..{to_str} top_keys={top_keys} "
        f"use_days_blocks={n_blocks} paid_log_count={len(paid_snapshots)} paid_logs={payload}"
    )


def _extract_paid_use_log_start_end(log: dict) -> tuple[object, object]:
    """
    Jobcan use-days の use_log から開始/終了時刻らしき値を抽出する。
    キー名揺れに備えて候補を順に見る。見つからなければ (None, None)。
    """
    if not isinstance(log, dict):
        return None, None

    containers = [
        log,
        log.get("use_days"),
        log.get("detail"),
        (log.get("detail") or {}).get("holiday") if isinstance(log.get("detail"), dict) else None,
    ]
    key_pairs = [
        ("start", "end"),
        ("start_time", "end_time"),
        ("from", "to"),
        ("from_time", "to_time"),
        ("start_at", "end_at"),
        ("use_start", "use_end"),
        ("use_start_time", "use_end_time"),
        ("use_from", "use_to"),
    ]

    for c in containers:
        if not isinstance(c, dict):
            continue
        for k1, k2 in key_pairs:
            st = c.get(k1)
            et = c.get(k2)
            if st and et:
                return st, et
    return None, None


def _infer_net_paid_leave_anchor_start(log: dict) -> str:
    """
    ネット事業部の半休で開始/終了が取得できない場合のフォールバック開始時刻を推定する。
    Jobcan の use-days には「午後」等の文言が入ることがあるため、文字列化して判定する。
    """
    try:
        raw = json.dumps(log, ensure_ascii=False, default=str)
    except Exception:
        raw = str(log)
    if "午後" in raw or "PM" in raw or "pm" in raw:
        return "13:00"
    return "09:00"


def _normalize_hhmm_pair(start_raw, end_raw) -> tuple[object, object]:
    """
    開始/終了時刻候補を HH:MM 文字列に正規化する。
    不正な場合は (None, None)。
    """
    if not isinstance(start_raw, str) or not isinstance(end_raw, str):
        return None, None
    start_s = start_raw.strip()
    end_s = end_raw.strip()
    if not start_s or not end_s:
        return None, None
    try:
        datetime.strptime(start_s, "%H:%M")
        datetime.strptime(end_s, "%H:%M")
    except ValueError:
        return None, None
    return start_s, end_s


def _load_work_kind_profile(work_kind_id):
    """
    users.work_kind_id をキーに work_kind ドキュメントを読み、minutes/start/end を返す。
    不正・未取得は各値 None。
    """
    profile = {"work_kind_id": work_kind_id, "minutes": None, "start": None, "end": None}
    if work_kind_id is None:
        return profile

    try:
        wk_doc = db.collection("work_kind").document(str(work_kind_id).strip()).get()
        if not wk_doc.exists:
            return profile
        d = wk_doc.to_dict() or {}
        try:
            m = d.get("minutes")
            if m is not None:
                mv = int(float(m))
                if mv >= 0:
                    profile["minutes"] = mv
        except (TypeError, ValueError):
            pass
        st, et = _normalize_hhmm_pair(d.get("start"), d.get("end"))
        profile["start"], profile["end"] = st, et
    except Exception as e:
        current_app.logger.warning(f"Failed to load work_kind profile for work_kind_id={work_kind_id}: {e}")
    return profile


@api_bp.route("/sync-paid-holidays", methods=["POST"])
@token_required
def sync_paid_holidays():
    """
    指定された月度の勤怠情報（有休・宿泊備考など）をJobcanから取得し、日報として反映する。
    """
    user_id = g.line_user_id
    data = request.get_json()
    target_date_str = data.get('date') # YYYY-MM-DD (月度内の任意の日付)
    target_employee_id_req = data.get('target_employee_id') # 管理者用: 対象従業員ID

    if not target_date_str:
        abort(400, "Request body must contain 'date'.")

    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
    except ValueError:
        abort(400, "Invalid date format.")

    # 1. ユーザー情報の取得
    user_info = get_user_info_by_line_id(user_id)
    
    # デフォルトは実行者本人
    target_company_id = user_info.get("company_employee_id")
    target_jobcan_id = user_info.get("jobcan_employee_id")

    # 管理者が他人のIDを指定した場合
    mapping_doc = None
    if target_employee_id_req:
        if not user_info.get('is_manager'):
            abort(403, "管理者権限が必要です。")
        
        target_company_id = target_employee_id_req
        # マッピングからJobcan IDを取得
        mapping_doc = db.collection("employee_mappings").document(target_company_id).get()
        if not mapping_doc.exists:
             abort(404, "指定された従業員が見つかりません。")
        target_jobcan_id = mapping_doc.to_dict().get("jobcan_employee_id")
    else:
        # 本人の場合は employee_mappings から status を取得するため取得
        mapping_doc = db.collection("employee_mappings").document(target_company_id).get()

    mapping_status = mapping_doc.to_dict().get("status") if mapping_doc and mapping_doc.exists else None

    # 2. 期間の計算 (月度)
    start_date, end_date = calculate_monthly_period(target_date)
    from_str = start_date.strftime('%Y-%m-%d')
    to_str = end_date.strftime('%Y-%m-%d')

    # Jobcan ID が実質ない（None/空文字のみ）、または status が active_officer の場合は
    # JOBCAN にIDが無い前提のため、エラーにせず当月の全日付で勤務時間を 0 に設定して成功で返す。
    # 0 や '0' は「IDあり」と扱い誤ゼロ化を防ぐ。
    if _is_effectively_no_jobcan_id(target_jobcan_id) or mapping_status == "active_officer":
        updated_count = 0
        current = start_date
        while current <= end_date:
            date_str = current.strftime('%Y-%m-%d')
            doc_id = f"{target_company_id}_{date_str}"
            doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
            doc = doc_ref.get()
            if doc.exists:
                doc_ref.update({"jobcan_work_minutes": 0})
                updated_count += 1
            current += timedelta(days=1)
        current_app.logger.info(
            f"[sync_paid_holidays] No Jobcan ID or active_officer for {target_company_id}. "
            f"Set jobcan_work_minutes=0 for {updated_count} existing report(s) in period {from_str}–{to_str}."
        )
        # フロントエンドは result.count を参照しているため、従来のフィールド名に合わせて count も返す
        return jsonify({
            "status": "success",
            "message": "Jobcan IDがありません（active_officer 等）。当月の既存日報の勤務時間を0分に設定しました。",
            "updated_count": updated_count,
            "count": updated_count
        }), 200

    # 3. Jobcanからデータ取得
    from services.jobcan_service import JobcanService
    app_env = os.environ.get("APP_ENV", "development")
    is_sandbox = app_env != "production"
    
    jobcan_service = JobcanService(
        db=db,
        sandbox=is_sandbox,
        raw_responses_collection=COLLECTION_JOBCAN_RAW_RESPONSES
    )
    
    processed_count = 0

    # 同期対象ユーザーの main_group / work_kind_id
    # - main_group: ネット事業部判定
    # - work_kind_id: use_id=1（全日有休）での minutes / 時刻補完
    target_user_snap = db.collection("users").document(target_company_id).get()
    target_main_group = None
    target_work_kind_id = None
    if target_user_snap.exists:
        target_user_data = target_user_snap.to_dict() or {}
        target_main_group = target_user_data.get("main_group")
        target_work_kind_id = target_user_data.get("work_kind_id")
    is_net_sync_target = is_net_main_group(target_main_group)
    work_kind_profile = _load_work_kind_profile(target_work_kind_id)

    # --- デバッグ用: 同期待ち対象の概要をログ出力 ---
    current_app.logger.info(
        f"[sync_paid_holidays] target_company_id={target_company_id}, "
        f"jobcan_employee_id={target_jobcan_id}, period={from_str}–{to_str}, "
        f"mapping_status={mapping_status}, work_kind_id={target_work_kind_id}, main_group={target_main_group}"
    )

    # --- A. 有休情報の取得と反映 ---
    holidays_data = jobcan_service.get_paid_holidays(target_jobcan_id, from_str, to_str, "paid")

    # TEMP: Jobcan レスポンス構造の確認（Cloud Run ログ）
    _temp_log_jobcan_paid_use_days_response(
        current_app.logger,
        holidays_data,
        str(target_company_id),
        target_jobcan_id,
        from_str,
        to_str,
    )

    # --- デバッグ用: use_days の要約をログ出力（対象ユーザー・期間のみに限定） ---
    if holidays_data and holidays_data.get("use_days"):
        try:
            summary_logs = []
            for emp in holidays_data.get("use_days", []):
                for log in emp.get("use_logs", []):
                    detail = log.get("detail", {}) or {}
                    use_days_obj = log.get("use_days", {}) or {}
                    summary_logs.append({
                        "use_date": log.get("use_date"),
                        "type": detail.get("type"),
                        "use_id": log.get("use_id"),
                        "days": use_days_obj.get("days"),
                    })
            current_app.logger.info(
                f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} use_logs_summary "
                f"company_emp={target_company_id} jobcan_id={target_jobcan_id} "
                f"count={len(summary_logs)} logs={summary_logs}"
            )
        except Exception as e:
            current_app.logger.warning(
                f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} failed to summarize use_logs for logging: {e}"
            )

    if holidays_data and holidays_data.get("use_days"):
        for employee_data in holidays_data["use_days"]:
             if employee_data.get("use_logs"):
                for log in employee_data["use_logs"]:
                    # detail.type == 'paid' チェック
                    if log.get("detail", {}).get("type") == "paid":
                        use_date_str = log.get("use_date")
                        if not use_date_str: continue
                        
                        use_date = datetime.strptime(use_date_str, '%Y-%m-%d')
                        
                        # 全休/半休判定（use_days.days）
                        days = 0
                        if log.get("use_days"):
                             days = float(log["use_days"].get("days", 0))
                        
                        holiday_type = None
                        if days >= 1.0:
                            holiday_type = 'full'
                        elif days > 0:
                            holiday_type = 'half'
                        
                        if holiday_type:
                            # デバッグ用: 判定直後の情報をログ出力（minutes 決定前）
                            current_app.logger.info(
                                f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} candidate "
                                f"company_emp={target_company_id} date={use_date_str} "
                                f"use_id={log.get('use_id')!r} days={days} "
                                f"holiday_bucket={holiday_type} work_kind_id={target_work_kind_id!r}"
                            )

                            # 工務: holiday_types.minutes（Firestore）を優先。use_id は work_kind_id または
                            # holiday_type_id（doc id）と照合。無ければ 480/240。
                            # ネット事業部: 同一照合で start/end を取り、タスクに startTime/endTime を付与（/api/reports_net と整合）。
                            use_id = log.get("use_id")
                            if is_net_sync_target:
                                # half は target_work_kind_id 紐づきの holiday_types を優先。
                                # full は従来どおり use_id ベースで解決する。
                                if holiday_type == "half" and target_work_kind_id is not None:
                                    pl = resolve_paid_leave_for_sync_by_work_kind(target_work_kind_id, holiday_type)
                                else:
                                    pl = resolve_paid_leave_for_sync(use_id, holiday_type)
                                minutes = pl["minutes"]
                                minutes_src = "holiday_types_or_fallback"
                                # use_id=1 の work_kind.minutes 優先は full の場合のみ適用。
                                # half は holiday_types（または fallback=240）を優先する。
                                if (
                                    str(use_id).strip() == "1"
                                    and holiday_type == "full"
                                    and work_kind_profile.get("minutes") is not None
                                ):
                                    minutes = int(work_kind_profile.get("minutes"))
                                    minutes_src = "work_kind.minutes"

                                # ネットの時刻は use-days レスポンスを最優先
                                st_raw, et_raw = _extract_paid_use_log_start_end(log)
                                st, et = _normalize_hhmm_pair(st_raw, et_raw)
                                time_src = "use_days_log"

                                # use-days に時刻が無い場合の補完順:
                                # - half: holiday_types.holiday を優先
                                # - full: use_id=1 の場合は work_kind.start/end を優先
                                if not st or not et:
                                    if holiday_type == "half":
                                        st, et = _normalize_hhmm_pair(pl.get("startTime"), pl.get("endTime"))
                                        if st and et:
                                            time_src = "holiday_types.holiday"
                                    else:
                                        if str(use_id).strip() == "1":
                                            st, et = _normalize_hhmm_pair(
                                                work_kind_profile.get("start"),
                                                work_kind_profile.get("end"),
                                            )
                                            if st and et:
                                                time_src = "work_kind.start_end"

                                # 既存挙動の保険: まだ未決定なら残り候補を試す
                                if not st or not et:
                                    if holiday_type == "half":
                                        if str(use_id).strip() == "1":
                                            st, et = _normalize_hhmm_pair(
                                                work_kind_profile.get("start"),
                                                work_kind_profile.get("end"),
                                            )
                                            if st and et:
                                                time_src = "work_kind.start_end"
                                    else:
                                        st, et = _normalize_hhmm_pair(pl.get("startTime"), pl.get("endTime"))
                                        if st and et:
                                            time_src = "holiday_types.holiday"
                                if not st or not et:
                                    anchor = _infer_net_paid_leave_anchor_start(log)
                                    st, et = default_net_paid_leave_time_slot(minutes, anchor_start=anchor)
                                    time_src = f"synthetic_{anchor.replace(':','')}h00_plus_minutes"
                                applied = register_paid_holiday_work_report(
                                    target_employee_id=target_company_id,
                                    target_date=use_date,
                                    minutes=minutes,
                                    inputter_info=user_info,
                                    start_time=st,
                                    end_time=et,
                                )
                                # TEMP: use_id → 採用した minutes / 時刻（Cloud Run ログ）
                                current_app.logger.info(
                                    f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} resolve: company_emp={target_company_id} "
                                    f"date={use_date_str} use_id={use_id!r} holiday_bucket={holiday_type} "
                                    f"minutes={minutes} net=True startTime={st!r} endTime={et!r} "
                                    f"minutes_source={minutes_src} time_source={time_src} "
                                    f"work_kind_id={target_work_kind_id!r} applied={applied}"
                                )
                            else:
                                minutes = resolve_paid_leave_minutes_engineering(use_id, holiday_type)
                                minutes_src = "holiday_types_or_fallback"
                                # use_id=1 の work_kind.minutes 優先は full の場合のみ適用。
                                # half は holiday_types（または fallback=240）を優先する。
                                if (
                                    str(use_id).strip() == "1"
                                    and holiday_type == "full"
                                    and work_kind_profile.get("minutes") is not None
                                ):
                                    minutes = int(work_kind_profile.get("minutes"))
                                    minutes_src = "work_kind.minutes"
                                applied = register_paid_holiday_work_report(
                                    target_employee_id=target_company_id,
                                    target_date=use_date,
                                    minutes=minutes,
                                    inputter_info=user_info,
                                )
                                current_app.logger.info(
                                    f"{_TEMP_PAID_HOLIDAY_SYNC_LOG_TAG} resolve: company_emp={target_company_id} "
                                    f"date={use_date_str} use_id={use_id!r} holiday_bucket={holiday_type} "
                                    f"minutes={minutes} net=False minutes_source={minutes_src} "
                                    f"work_kind_id={target_work_kind_id!r} applied={applied}"
                                )
                            if applied:
                                processed_count += 1

    # --- B. 選択備考（宿泊など）の取得と反映 ---
    try:
        selection_notes = jobcan_service.get_selection_notes(target_jobcan_id, from_str, to_str)
        current_app.logger.info(f"Syncing selection notes for {target_company_id}. Found {len(selection_notes)} notes.")
        
        for note_date_str, note_content in selection_notes.items():
            # get_selection_notesで既にID=1(宿泊)に絞り込んでいるため、データが存在すれば宿泊ありとみなす
            # note_contentには "1" (コード) や "宿泊" (名称) などが入る
            if note_content:
                doc_id = f"{target_company_id}_{note_date_str}"
                doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
                
                doc = doc_ref.get()
                if doc.exists:
                    doc_ref.update({
                        "has_accommodation": True,
                        "jobcan_note": note_content, # 念のため内容も保存
                        "report_updated_at": firestore.SERVER_TIMESTAMP
                    })
                else:
                    # 日報が存在しない場合、新規作成する
                    try:
                        note_date = datetime.strptime(note_date_str, '%Y-%m-%d')
                        doc_ref.set({
                            "company_employee_id": target_company_id,
                            "date": note_date,
                            "has_accommodation": True,
                            "jobcan_note": note_content,
                            "report_updated_at": firestore.SERVER_TIMESTAMP,
                            "task_total_minutes": 0 # 初期値
                        })
                    except ValueError:
                        current_app.logger.error(f"Invalid date format for note: {note_date_str}")
                
                # processed_count は有休と重複する可能性があるため、ここではインクリメントしないか、
                # 別途カウントするかは仕様次第です。

    except Exception as e:
        # 選択備考の取得失敗は全体のエラーにせず、ログ出力に留める
        current_app.logger.warning(f"Failed to sync selection notes: {e}")

    # --- C. 現場作業時間の集計とon_siteフラグの設定 ---
    try:
        # インデックス未作成エラーを回避するため、ドキュメントID指定で取得する
        dates_list = []
        curr = start_date
        while curr <= end_date:
            dates_list.append(curr)
            curr += timedelta(days=1)
            
        doc_refs = []
        for d in dates_list:
            d_str = d.strftime('%Y-%m-%d')
            doc_id = f"{target_company_id}_{d_str}"
            doc_refs.append(db.collection(COLLECTION_DAILY_REPORTS).document(doc_id))
            
        # get_all で一括取得 (存在しないドキュメントも含まれるため後でチェック)
        docs = db.get_all(doc_refs)
        
        batch = db.batch()
        update_count = 0
        for doc in docs:
            if not doc.exists:
                continue
                
            doc_data = doc.to_dict()
            tasks = doc_data.get("tasks", [])
            
            # A01 / A12（現場対象）の時間を集計 (念のためint変換して加算)
            on_site_category_ids = {"A01", "A12"}
            on_site_minutes = 0
            for task in tasks:
                if task.get("categoryA_id") in on_site_category_ids:
                    try:
                        on_site_minutes += int(task.get("time", 0))
                    except (ValueError, TypeError):
                        pass
            
            new_on_site_status = None
            if on_site_minutes >= 360:
                new_on_site_status = "full"
            elif on_site_minutes >= 240:
                new_on_site_status = "half"
            
            current_status = doc_data.get("on_site")
            
            if new_on_site_status != current_status:
                if new_on_site_status:
                    batch.update(doc.reference, {"on_site": new_on_site_status})
                else:
                    # 以前は設定されていたが、条件を満たさなくなった場合はフィールドを削除
                    if current_status:
                        batch.update(doc.reference, {"on_site": firestore.DELETE_FIELD})
                update_count += 1
        
        if update_count > 0:
            batch.commit()
            current_app.logger.info(f"Updated on_site status for {update_count} reports for user {target_company_id}.")
        else:
            current_app.logger.info(f"No on_site status updates needed for user {target_company_id}.")
            
    except Exception as e:
        current_app.logger.error(f"Failed to sync on_site status: {e}", exc_info=True)

    return jsonify({"status": "success", "count": processed_count}), 200

def _task_time_int_for_report(task: dict) -> int:
    """タスクの time を報告用に整数化する。"""
    t = task.get("time", 0)
    try:
        return int(float(t))
    except (TypeError, ValueError):
        return 0


def _format_report_content_line_from_task(task: dict) -> str:
    """
    業務内容テキストの1行。startTime/endTime がある場合は reports_net 形式、なければ工務形式。
    """
    st = task.get("startTime")
    et = task.get("endTime")
    if st and et:
        tm = _task_time_int_for_report(task)
        return (
            f"【{st}~{et}|{tm}分】"
            f"{task.get('categoryA_label', '')} - {task.get('categoryB_label', '')}"
        )
    detail = task.get("detail", "")
    tm = _task_time_int_for_report(task)
    return (
        f"[{task.get('categoryA_label', '')}/{task.get('categoryB_label', '')}] "
        f"{tm}分: {detail}"
    )


# --- 内部関数: 有休自動入力用 ---
def _task_looks_like_paid_leave(task: dict) -> bool:
    """
    日報 tasks の1要素が「有休」として既に存在するかの広めの判定。
    自動付与は A00/e_000000 だが、手入力・過去データは ID やラベルの組み合わせがずれることがある。
    いずれかに該当すれば Jobcan 同期では追加・置換しない（上書き防止）。
    """
    if not isinstance(task, dict):
        return False
    ca = str(task.get("categoryA_id") or "").strip()
    cl = str(task.get("categoryA_label") or "").strip()
    cb = str(task.get("categoryB_id") or "").strip()
    bl = str(task.get("categoryB_label") or "").strip()
    if ca == "A00":
        return True
    if ca == "A99":
        return True
    if cl == "有休":
        return True
    if cl == "有休(忌引)":
        return True
    if cb == "e_000000":
        return True
    if bl == "000000" and (cl == "有休" or cl == "有休(忌引)" or ca == "A00" or ca == "A99"):
        return True
    return False


def register_paid_holiday_work_report(
    target_employee_id,
    target_date,
    minutes,
    inputter_info,
    start_time=None,
    end_time=None,
):
    """
    有休情報を元に日報データを生成・更新する関数。
    戻り値: 実際に作成/更新した場合は True、既存有休タスクがありスキップした場合は False。

    既存ドキュメントに有休らしきタスクが1件でもあれば何もしない（tasks を書き換えない）。
    
    Args:
        target_employee_id (str): 対象従業員の社内ID
        target_date (datetime): 対象日 (datetimeオブジェクト)
        minutes (int): 有休タスクの分数（工務: holiday_types.minutes または 480/240 フォールバック）
        inputter_info (dict): 実行者のユーザー情報 (company_employee_id, nameを含む)
        start_time (str|None): ネット事業部向け。開始時刻（HH:MM）。end_time とセットで指定。
        end_time (str|None): ネット事業部向け。終了時刻（HH:MM）。
    """
    try:
        minutes = int(minutes)
    except (TypeError, ValueError):
        minutes = 0
    if minutes < 0:
        minutes = 0

    CATEGORY_A_ID = "A00"
    CATEGORY_A_LABEL = "有休"
    CATEGORY_B_ID = "e_000000"
    CATEGORY_B_LABEL = "000000"

    use_net_task_shape = bool(start_time and end_time)

    # タスクオブジェクト
    new_task = {
        "categoryA_id": CATEGORY_A_ID,
        "categoryA_label": CATEGORY_A_LABEL,
        "categoryB_id": CATEGORY_B_ID,
        "categoryB_label": CATEGORY_B_LABEL,
        "time": minutes,
        "detail": "有休自動適用",
    }
    if use_net_task_shape:
        new_task["startTime"] = start_time
        new_task["endTime"] = end_time
        new_task["comment"] = ""

    doc_id = f"{target_employee_id}_{target_date.strftime('%Y-%m-%d')}"
    doc_ref = db.collection(COLLECTION_DAILY_REPORTS).document(doc_id)
    
    try:
        doc = doc_ref.get()
        
        if doc.exists:
            # 既存データがある場合
            data = doc.to_dict()
            tasks = data.get('tasks', [])

            paid_task_indexes = [idx for idx, t in enumerate(tasks) if _task_looks_like_paid_leave(t)]
            has_existing_paid_task = len(paid_task_indexes) > 0

            if has_existing_paid_task:
                # 要件: 既に有休らしきタスクがあれば追加も置換もしない（Jobcan 同期の二重適用・上書き防止）
                current_app.logger.info(
                    f"Skipped paid holiday sync for {target_employee_id} on {target_date.strftime('%Y-%m-%d')} "
                    f"because paid-leave-like task(s) already exist (indices={paid_task_indexes})."
                )
                return False

            # 有休タスクが未登録の場合のみ追加
            tasks.append(new_task)
            
            # 合計時間の再計算
            new_total = 0
            for t in tasks:
                new_total += _task_time_int_for_report(t)
            
            # 業務内容テキストの更新（ネット形式・工務形式が混在しても行ごとに整形）
            task_details = [_format_report_content_line_from_task(t) for t in tasks]
            new_report_content = "\n".join(task_details)

            doc_ref.update({
                "tasks": tasks,
                "task_total_minutes": new_total,
                "report_content": new_report_content,
                "report_updated_at": firestore.SERVER_TIMESTAMP
            })
            current_app.logger.info(f"Updated paid holiday task for {target_employee_id} on {target_date.strftime('%Y-%m-%d')}")
            return True

        else:
            # 新規作成
            # ターゲットユーザーの詳細情報を取得 (グループ情報など)
            target_user_ref = db.collection("users").document(target_employee_id)
            target_user_doc = target_user_ref.get()
            
            target_group_id = None
            target_group_name = None
            target_employee_name = "Unknown"

            if target_user_doc.exists:
                user_data = target_user_doc.to_dict()
                target_group_id = user_data.get("main_group")
                # グループ名取得
                if target_group_id:
                    group_doc = db.collection("group_mappings").document(str(target_group_id)).get()
                    if group_doc.exists:
                        target_group_name = group_doc.to_dict().get("name")
            
            # 名前取得
            mapping_doc = db.collection("employee_mappings").document(target_employee_id).get()
            if mapping_doc.exists:
                target_employee_name = mapping_doc.to_dict().get("name")

            # レポート内容テキスト（ネットは reports_net と同じ行形式）
            if use_net_task_shape:
                report_content = (
                    f"【{start_time}~{end_time}|{minutes}分】"
                    f"{CATEGORY_A_LABEL} - {CATEGORY_B_LABEL}"
                )
            else:
                report_content = f"[{CATEGORY_A_LABEL}/{CATEGORY_B_LABEL}] {minutes}分: 有休自動適用"

            new_report_data = {
                "employee_name": target_employee_name,
                "inputter_id": inputter_info.get("company_employee_id"),
                "inputter_name": inputter_info.get("name"),
                "is_proxy_report": False,
                "group_id": target_group_id,
                "group_name": target_group_name,
                "report_year": target_date.year,
                "report_month": target_date.month,
                "jobcan_work_minutes": 0, # 新規作成時は0
                "date": target_date,
                "company_employee_id": target_employee_id,
                "task_total_minutes": minutes,
                "tasks": [new_task],
                "report_content": report_content,
                "report_updated_at": firestore.SERVER_TIMESTAMP
            }
            
            doc_ref.set(new_report_data)
            current_app.logger.info(f"Created new paid holiday report for {target_employee_id} on {target_date.strftime('%Y-%m-%d')}")
            return True

    except Exception as e:
        current_app.logger.error(f"Failed to register paid holiday report: {e}")
        raise e